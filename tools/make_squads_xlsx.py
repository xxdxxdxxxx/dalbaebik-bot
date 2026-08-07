"""One-shot: create squads.xlsx from players.json (в корне проекта)."""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
DB = ROOT / "players.json"
OUT = ROOT / "squads.xlsx"

players: dict = {}
if DB.exists():
    data = json.loads(DB.read_text(encoding="utf-8"))
    players = data.get("players") or {}

wb = Workbook()
ws = wb.active
ws.title = "roster"

headers = ["squad", "slot", "discord_id", "game_nick", "discord_name", "note"]
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True)
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt = PatternFill("solid", fgColor="E8F0FE")

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

SQUADS, SLOTS = 6, 5
plist = sorted(players.items(), key=lambda x: (x[1].get("game_nick") or "").lower())
idx = 0
row = 2
for squad in range(1, SQUADS + 1):
    for slot in range(1, SLOTS + 1):
        did, p = (plist[idx] if idx < len(plist) else (None, {}))
        if did is not None:
            idx += 1
        values = [
            squad,
            slot,
            # ВАЖНО: discord_id только СТРОКОЙ. Числом Excel ломает snowflake (>2^53).
            str(did) if did else "",
            (p.get("game_nick") or "") if did else "",
            (p.get("discord_name") or p.get("discord_username") or "") if did else "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col, val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col <= 3 else "left")
            if col == 3:
                cell.number_format = "@"  # текст
            if squad % 2 == 0:
                cell.fill = alt
        row += 1

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 24

ws2 = wb.create_sheet("squads")
ws2["A1"] = "squad"
ws2["B1"] = "name"
ws2["A1"].font = header_font
ws2["B1"].font = header_font
ws2["A1"].fill = header_fill
ws2["B1"].fill = header_fill
names = ["Альфа", "Браво", "Чарли", "Дельта", "Эхо", "Фокстрот"]
for i, name in enumerate(names, 1):
    ws2.cell(i + 1, 1, i)
    ws2.cell(i + 1, 2, name)
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 16

ws3 = wb.create_sheet("README")
ws3["A1"] = "Как пользоваться"
ws3["A1"].font = Font(bold=True, size=14)
help_lines = [
    "1. Лист roster: строка = слот (по умолчанию 6 отрядов x 5 человек).",
    "2. discord_id обязателен и ТОЛЬКО ТЕКСТОМ (перед вставкой: ' или формат ячейки Текст). Иначе Excel портит ID.",
    "3. game_nick — ник STALCRAFT. Меняешь здесь → бот обновит явку и гранаты.",
    "4. squad / slot — отряд и место. Чтобы поменять людей — поменяй ячейки между строками.",
    "5. Пустой discord_id = пустой слот.",
    "6. Лист squads — имена отрядов.",
    "7. Сохрани файл. Бот читает раз в минуту; /sheet_sync — сразу.",
    "8. Не удаляй первую строку-заголовок roster.",
    "9. Для правок с сайта: положи файл в OneDrive и укажи SHEET_PATH в .env на локальный путь.",
    "10. Если Excel на ПК держит файл — иногда PermissionError; Excel Online обычно ок.",
]
for i, line in enumerate(help_lines, 3):
    ws3.cell(i, 1, line)
ws3.column_dimensions["A"].width = 110

wb.save(OUT)
print(f"saved {OUT.resolve()}")
print(f"placed={min(len(plist), SQUADS * SLOTS)} total_players={len(plist)}")
