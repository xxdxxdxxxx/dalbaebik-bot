"""
Discord-бот для КВ STALCRAFT:
- база Discord ↔ ник в игре (/add, /remove)
- явка: чт–сб 19:30 МСК · вс 18:30 МСК
- грены по этапам (разное расписание чт–сб / вс)
- карты этапов (/map, !map) → embed ГРАНАТЫ + txt-сканы
- отряды + синк из Excel (SHEET_PATH)
"""

from __future__ import annotations

import asyncio
import json
import os
import random
import re
import time
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import Any
from urllib.parse import quote

import aiohttp
import discord
import pytz
from discord import app_commands
from discord.ext import commands, tasks
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# ---------------------------------------------------------------------------
# Логи в консоль — коротко и по делу
# ---------------------------------------------------------------------------
def log(msg: str, level: str = "info") -> None:
    """level: info | ok | warn | err | kv"""
    tag = {
        "info": "·",
        "ok": "✓",
        "warn": "!",
        "err": "✗",
        "kv": "◆",
    }.get(level, "·")
    print(f"  {tag}  {msg}")


def log_banner(lines: list[str]) -> None:
    w = 46
    print()
    print("  ╔" + "═" * w + "╗")
    for line in lines:
        s = line if len(line) <= w - 2 else line[: w - 3] + "…"
        print(f"  ║ {s:<{w - 2}} ║")
    print("  ╚" + "═" * w + "╝")
    print()


# ---------------------------------------------------------------------------
# Настройки (.env — только токен и API; каналы задаются командами в Discord)
# ---------------------------------------------------------------------------
DISCORD_TOKEN = os.getenv("DISCORD_TOKEN", "")
# Опциональный запасной вариант из .env (если не настроили через /setup)
ENV_GUILD_ID = int(os.getenv("GUILD_ID", "0") or 0)
ENV_LOG_CHANNEL_ID = int(os.getenv("LOG_CHANNEL_ID", "0") or 0)
ENV_VOICE_CHANNEL_IDS = [
    int(x.strip())
    for x in os.getenv("VOICE_CHANNEL_IDS", "").split(",")
    if x.strip().isdigit()
]

CLIENT_ID = os.getenv("STALCRAFT_CLIENT_ID", "941")
CLIENT_SECRET = os.getenv("STALCRAFT_CLIENT_SECRET", "")
# Регион API: только RU-сервер (eapi …/ru/…)
REGION = (os.getenv("STALCRAFT_REGION", "ru") or "ru").strip().lower()
if REGION not in ("ru", "eu", "na", "sea", "nea"):
    log(f"REGION={REGION!r} → ru", "warn")
    REGION = "ru"

# ---------------------------------------------------------------------------
# Пути проекта
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
PHRASES_DIR = ROOT / "phrases"
SCANS_DIR = ROOT / "scans"
SCANS_ETAPY_DIR = SCANS_DIR / "etapy"   # база / этап1 / этап2 / финал
SCANS_ITOGI_DIR = SCANS_DIR / "itogi"   # только итоги дня

# Excel-таблица отрядов (OneDrive-путь или файл в корне проекта)
_default_sheet = ROOT / "ДИТЯ22.xlsx"
if not _default_sheet.exists():
    _default_sheet = ROOT / "squads.xlsx"
SHEET_PATH = Path(os.getenv("SHEET_PATH", str(_default_sheet)).strip() or str(_default_sheet))
SHEET_SYNC_SECONDS = max(15, int(os.getenv("SHEET_SYNC_SECONDS", "60") or 60))

MSK = pytz.timezone("Europe/Moscow")
DB_PATH = ROOT / "players.json"
# Одна строка = одно сообщение; после !add / /add шлётся случайная через ~5 сек
ADD_PHRASES_FILE = PHRASES_DIR / "add_phrases.txt"
# ЛС тем, кто ни разу не зашёл в войс (авто 19:50 + /dm_absent)
ABSENT_DM_PHRASES_FILE = PHRASES_DIR / "absent_dm_phrases.txt"

# slug этапа → имя файла скана (внутренние id шагов)
GRENADE_STEP_SLUG = {
    "20:00": "база",
    "20:25": "этап1",
    "20:50": "этап2",
    "21:20": "этап3",
    "21:40": "этап4",  # только вс (финал)
}

# Дни КВ: чт / пт / сб / вс (пн=0 … вс=6)
CW_WEEKDAYS = (3, 4, 5, 6)

# Отряд «Чемпионы» / замены (Excel F6:F16) — не пинговать ЛС перед КВ
SQUAD_SUBS_ID = 99

# ---------------------------------------------------------------------------
# Расписание КВ по дням (МСК)
# чт–сб: 3 этапа — база 20:05, I 20:25, II 20:50, III 21:20
# вс:    4 этапа —
#        19:00–19:20 · 19:20–19:40 · 19:40–20:00 · 20:00–20:20
#        сканы: 19:00 база, 19:20 I, 19:40 II, 20:00 III, 20:20 IV финал
# ---------------------------------------------------------------------------
def _weekday(dt: datetime | None = None) -> int:
    if dt is None:
        dt = datetime.now(MSK)
    return dt.weekday()


def kv_start(dt: datetime | None = None) -> dt_time:
    """Начало явки."""
    return dt_time(18, 30) if _weekday(dt) == 6 else dt_time(19, 30)


def kv_absent_dm(dt: datetime | None = None) -> dt_time:
    """ЛС неявившимся."""
    return dt_time(18, 45) if _weekday(dt) == 6 else dt_time(19, 50)


def kv_attendance_end(dt: datetime | None = None) -> dt_time:
    """Конец сбора ✅ (дальше только 🔊)."""
    return dt_time(19, 0) if _weekday(dt) == 6 else dt_time(20, 5)


def kv_grenade_steps(dt: datetime | None = None) -> list[tuple[str, dt_time, str | None]]:
    """(step_id, wall_clock, prev_step_id). step_id стабилен для history/embed."""
    if _weekday(dt) == 6:
        return [
            ("20:00", dt_time(19, 0), None),       # база / старт I
            ("20:25", dt_time(19, 20), "20:00"),   # конец I  19:00–19:20
            ("20:50", dt_time(19, 40), "20:25"),   # конец II 19:20–19:40
            ("21:20", dt_time(20, 0), "20:50"),    # конец III 19:40–20:00
            ("21:40", dt_time(20, 20), "21:20"),   # конец IV 20:00–20:20 финал
        ]
    return [
        ("20:00", dt_time(20, 5), None),
        ("20:25", dt_time(20, 25), "20:00"),
        ("20:50", dt_time(20, 50), "20:25"),
        ("21:20", dt_time(21, 20), "20:50"),
    ]


def player_squad_id(p: dict[str, Any] | None) -> int | None:
    if not p:
        return None
    sq = p.get("squad")
    if sq is None or sq == "":
        return None
    try:
        return int(sq)
    except (TypeError, ValueError):
        return None


def is_subs_player(p: dict[str, Any] | None) -> bool:
    """Чемпионы / замены (squad 99) — без ЛС-пинга перед КВ."""
    return player_squad_id(p) == SQUAD_SUBS_ID


def is_scan_roster_player(p: dict[str, Any] | None) -> bool:
    """
    Кого сканим по API (только gre-thr):
    отряды 1–6 и Чемпионы (99). «Без отряда» — не сканим.
    Не зависит от войса / ✅.
    """
    sid = player_squad_id(p)
    if sid is None:
        return False
    if sid == SQUAD_SUBS_ID:
        return True
    return 1 <= sid <= 6


def kv_final_time(dt: datetime | None = None) -> dt_time:
    return kv_grenade_steps(dt)[-1][1]


def kv_final_step_id(dt: datetime | None = None) -> str:
    return kv_grenade_steps(dt)[-1][0]


def kv_step_wall_times(dt: datetime | None = None) -> dict[str, str]:
    """step_id → 'HH:MM' для подписей в логах."""
    return {sid: tm.strftime("%H:%M") for sid, tm, _ in kv_grenade_steps(dt)}


# совместимость: константы = расписание чт–сб (дефолт)
START_TIME = dt_time(19, 30)
ABSENT_DM_TIME = dt_time(19, 50)
ATTENDANCE_END = dt_time(20, 5)
GRENADE_STEPS = kv_grenade_steps(datetime(2024, 1, 4))  # четверг dummy

WEEKDAYS_RU = (
    "понедельник",
    "вторник",
    "среда",
    "четверг",
    "пятница",
    "суббота",
    "воскресенье",
)

SEP_LINE = "───────────────────────────────"

# Цвета оформления
COLOR_ONLINE = 0x2ECC71       # зелёный
COLOR_WAIT = 0x95A5A6         # серый — до сессии
COLOR_GRENADES = 0xE67E22     # оранжевый
COLOR_LIVE = 0xF1C40F         # жёлтый — live-отчёт
COLOR_OK = 0x57F287           # успех
COLOR_ERR = 0xED4245          # ошибка
COLOR_INFO = 0x5865F2         # blurple

STEP_TITLES = {
    "20:00": "База (старт)",
    "20:25": "Этап I",
    "20:50": "Этап II",
    "21:20": "Этап III",
    "21:40": "Этап IV (финал)",
}
# полный порядок (вс); чт–сб заканчиваются на 21:20
STEP_ORDER = ["20:00", "20:25", "20:50", "21:20", "21:40"]

# step_id → индекс карты (0=I, 1=II, 2=III, 3=IV); база без карты
STEP_MAP_INDEX = {
    "20:00": None,
    "20:25": 0,
    "20:50": 1,
    "21:20": 2,
    "21:40": 3,
}
STAGE_ROMAN = ("I", "II", "III", "IV")

# Канонические карты КВ + алиасы (нижний регистр)
MAP_ALIASES: dict[str, tuple[str, ...]] = {
    "Хвойник": (
        "хвойник", "хвоя", "хвой", "hvoynik", "hvoyn", "hvoy",
    ),
    "Бердовка": (
        "бердовка", "берда", "берд",
        "бердовка левая", "берда левая", "бердовка л", "берда л",
        "бердовка правая", "берда правая", "бердовка п", "берда п",
        "берд левая", "берд л", "берд правая", "берд п",
        "левая бердовка", "левая берда", "правая бердовка", "правая берда",
        "berda", "berdovka", "berd",
    ),
    "Низина": (
        "низина", "низ", "nizina", "niz",
    ),
}

# ---------------------------------------------------------------------------
# База данных
# ---------------------------------------------------------------------------
db_lock = asyncio.Lock()


def default_db() -> dict[str, Any]:
    return {
        "players": {},
        "message_ids": {"online": None, "grenades": None},
        "session_date": None,       # YYYY-MM-DD день текущей/последней сессии
        "grenade_date": None,       # дата для подписи под ГРАНАТЫ
        "grenade_history": {},
        # step_id пропущенных mid-этапов (заполнены forward-copy, дельта 0)
        "skipped_grenade_steps": [],
        "last_grenade_step": None,
        "kv_finished": False,       # после 3 этапа до следующего 19:30
        "kv_session_active": False, # True только после реального старта окна КВ (19:30)
        # discord_id str — уже слали ЛС «не был в войсе» в этой сессии (один раз)
        "absent_dm_sent": [],
        # id отряда (str) → название; подтягивается из Excel лист squads
        "squad_names": {},
        # карты дня: {"date": "YYYY-MM-DD", "maps": ["Хвойник", "Бердовка", "Низина"]}
        "kv_maps": {"date": None, "maps": []},
        "config": {
            "guild_id": None,
            "log_channel_id": None,
            "voice_channel_ids": [],
            # роли с полным правом /add /remove (как админ)
            "access_role_ids": [],
        },
    }


def load_db() -> dict[str, Any]:
    if not DB_PATH.exists():
        data = default_db()
        save_db(data)
        return data
    with open(DB_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    base = default_db()
    for k, v in base.items():
        if k not in data:
            data[k] = v
    if "message_ids" not in data or not isinstance(data["message_ids"], dict):
        data["message_ids"] = {"online": None, "grenades": None}
    if "config" not in data or not isinstance(data["config"], dict):
        data["config"] = default_db()["config"]
    else:
        for ck, cv in default_db()["config"].items():
            if ck not in data["config"]:
                data["config"][ck] = cv
    # HS% убран — не тащим combat_history
    data.pop("combat_history", None)
    return data


def save_db(data: dict[str, Any]) -> None:
    """Атомарная запись: temp → replace (не портит players.json при kill)."""
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = DB_PATH.with_name(DB_PATH.name + ".tmp")
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    with open(tmp, "w", encoding="utf-8") as f:
        f.write(payload)
        f.flush()
        try:
            os.fsync(f.fileno())
        except OSError:
            pass
    os.replace(tmp, DB_PATH)


def backup_db() -> None:
    """Копия players.json.bak (перед wipe сессии / reset)."""
    if not DB_PATH.exists():
        return
    bak = DB_PATH.with_name("players.json.bak")
    try:
        import shutil

        shutil.copy2(DB_PATH, bak)
    except OSError as e:
        log(f"backup db: {e}", "warn")


def _cfg_from_discord(cfg: dict[str, Any]) -> bool:
    """True, если каналы уже настраивали командами в Discord."""
    return bool(cfg.get("guild_id") or cfg.get("log_channel_id") or cfg.get("voice_channel_ids"))


def get_guild_id(data: dict[str, Any] | None = None) -> int:
    if data is None:
        data = load_db()
    cfg = data.get("config") or {}
    gid = cfg.get("guild_id")
    if gid:
        return int(gid)
    return ENV_GUILD_ID


def get_log_channel_id(data: dict[str, Any] | None = None) -> int:
    if data is None:
        data = load_db()
    cfg = data.get("config") or {}
    lid = cfg.get("log_channel_id")
    if lid:
        return int(lid)
    if _cfg_from_discord(cfg):
        return 0
    return ENV_LOG_CHANNEL_ID


def get_voice_channel_ids(data: dict[str, Any] | None = None) -> list[int]:
    if data is None:
        data = load_db()
    cfg = data.get("config") or {}
    ids = cfg.get("voice_channel_ids") or []
    if ids:
        return [int(x) for x in ids]
    # Уже настраивали в Discord (в т.ч. очистили список) — не подмешивать .env
    if _cfg_from_discord(cfg):
        return []
    return list(ENV_VOICE_CHANNEL_IDS)


def ensure_config(data: dict[str, Any]) -> dict[str, Any]:
    if "config" not in data or not isinstance(data["config"], dict):
        data["config"] = default_db()["config"]
    cfg = data["config"]
    if "access_role_ids" not in cfg or not isinstance(cfg.get("access_role_ids"), list):
        cfg["access_role_ids"] = []
    return cfg


def is_bot_admin(member: discord.Member | None) -> bool:
    """Админ сервера / Manage Guild — полный доступ к командам."""
    if member is None:
        return False
    perms = getattr(member, "guild_permissions", None)
    if perms is None:
        return False
    return bool(perms.administrator or perms.manage_guild)


def get_access_role_ids(data: dict[str, Any] | None = None) -> list[int]:
    if data is None:
        data = load_db()
    cfg = ensure_config(data)
    out: list[int] = []
    for x in cfg.get("access_role_ids") or []:
        try:
            out.append(int(x))
        except (TypeError, ValueError):
            continue
    return out


def member_role_ids(member: discord.Member) -> set[int]:
    """Все id ролей участника (в т.ч. из raw _roles)."""
    ids: set[int] = set()
    try:
        ids.update(int(r.id) for r in member.roles)
    except Exception:
        pass
    raw = getattr(member, "_roles", None)
    if raw is not None:
        try:
            ids.update(int(x) for x in raw)
        except Exception:
            pass
    if member.guild is not None:
        ids.add(int(member.guild.default_role.id))
    return ids


def member_has_access_role(member: discord.Member | None, data: dict[str, Any] | None = None) -> bool:
    if member is None:
        return False
    allowed = set(get_access_role_ids(data))
    if not allowed:
        return False
    return bool(member_role_ids(member) & allowed)


def can_use_roster(member: discord.Member | None, data: dict[str, Any] | None = None) -> bool:
    """Может вызывать /add и /remove."""
    if is_bot_admin(member):
        return True
    return member_has_access_role(member, data)


def can_manage_kv(member: discord.Member | None, data: dict[str, Any] | None = None) -> bool:
    """Скан / deletegren / reset_session / опасные KV-команды."""
    return can_use_roster(member, data)


def deny_embed(roles_hint: bool = True) -> discord.Embed:
    data = load_db() if roles_hint else None
    if data is not None:
        allowed = get_access_role_ids(data)
        roles_txt = ", ".join(f"<@&{r}>" for r in allowed) if allowed else "_ролей нет — только админ_"
        body = f"Нужна роль: {roles_txt}\nили админ сервера."
    else:
        body = "Недостаточно прав."
    return make_reply_embed("❌  Нет доступа", body, color=COLOR_ERR)


# ---------------------------------------------------------------------------
# Карты КВ (/map · !map)
# ---------------------------------------------------------------------------
def _norm_map_key(s: str) -> str:
    s = (s or "").strip().lower().replace("ё", "е")
    s = re.sub(r"[\s_\-]+", " ", s)
    return s.strip()


def _build_alias_lookup() -> list[tuple[str, str]]:
    """(alias_norm, canonical) — длинные алиасы первыми (жадный разбор)."""
    items: list[tuple[str, str]] = []
    for canon, aliases in MAP_ALIASES.items():
        for a in aliases:
            items.append((_norm_map_key(a), canon))
        items.append((_norm_map_key(canon), canon))
    items.sort(key=lambda x: len(x[0]), reverse=True)
    return items


_MAP_ALIAS_LOOKUP = _build_alias_lookup()


def canonicalize_map_name(raw: str) -> str:
    """Известный алиас → канон; иначе Title Case как ввёл пользователь."""
    key = _norm_map_key(raw)
    if not key:
        return ""
    for alias, canon in _MAP_ALIAS_LOOKUP:
        if key == alias:
            return canon
    # «бердалевая» без пробела и т.п.
    compact = key.replace(" ", "")
    for alias, canon in _MAP_ALIAS_LOOKUP:
        if compact == alias.replace(" ", ""):
            return canon
    return " ".join(w.capitalize() if w.isascii() else w for w in raw.strip().split())


def parse_maps_text(text: str, *, expect: int | None = None) -> tuple[list[str] | None, str | None]:
    """
    Разобрать строку карт.
    Разделители: | , ;  или жадный матч по каталогу.
    expect: сколько карт ждать (3 или 4); None = 3–4.
    """
    raw = (text or "").strip()
    if not raw:
        return None, "Укажи карты: `хвойник берда низина`"

    # явные разделители
    if any(sep in raw for sep in ("|", ",", ";")):
        parts = re.split(r"[|/;,]+", raw)
        maps = [canonicalize_map_name(p) for p in parts if p.strip()]
    else:
        # жадный разбор по известным картам
        rest = _norm_map_key(raw)
        maps = []
        original_rest = rest
        while rest:
            rest = rest.lstrip(" ")
            if not rest:
                break
            matched = False
            for alias, canon in _MAP_ALIAS_LOOKUP:
                if rest == alias or rest.startswith(alias + " "):
                    maps.append(canon)
                    rest = rest[len(alias):].lstrip(" ")
                    matched = True
                    break
            if not matched:
                # остаток одним куском (свободная карта) — если уже есть 2+ и остаток один токен-блок
                if maps and rest:
                    maps.append(canonicalize_map_name(rest))
                    rest = ""
                    break
                return None, (
                    f"Не разобрал карты: `{original_rest}`\n"
                    f"Известные: **хвойник**, **берда** / **бердовка**, **низина**\n"
                    f"Или через `|`: `хвойник | берда | низина`"
                )

    maps = [m for m in maps if m]
    if not maps:
        return None, "Не нашёл ни одной карты."

    if expect is not None:
        if len(maps) != expect:
            return None, f"Нужно **{expect}** карты(ы), а указано **{len(maps)}**."
    else:
        if len(maps) < 3:
            return None, f"Нужно минимум **3** карты (сейчас {len(maps)})."
        if len(maps) > 4:
            return None, f"Максимум **4** карты (вс), указано {len(maps)}."

    return maps, None


def get_kv_maps(data: dict[str, Any], *, for_date: str | None = None) -> list[str]:
    """Карты текущего дня КВ (пустой список, если не заданы / другой день)."""
    raw = data.get("kv_maps") or {}
    if not isinstance(raw, dict):
        return []
    day = for_date or data.get("grenade_date") or data.get("session_date") or today_msk_str()
    if raw.get("date") and raw.get("date") != day:
        return []
    maps = raw.get("maps") or []
    if not isinstance(maps, list):
        return []
    return [str(m).strip() for m in maps if str(m).strip()]


def set_kv_maps(data: dict[str, Any], maps: list[str], *, day: str | None = None) -> dict[str, Any]:
    data["kv_maps"] = {
        "date": day or today_msk_str(),
        "maps": list(maps),
    }
    return data


def clear_kv_maps(data: dict[str, Any]) -> dict[str, Any]:
    data["kv_maps"] = {"date": None, "maps": []}
    return data


def format_maps_lines(maps: list[str], *, max_stages: int | None = None) -> str:
    if not maps:
        return "_карты не заданы · `/map хвойник берда низина`_"
    n = max_stages if max_stages is not None else len(maps)
    lines: list[str] = []
    for i in range(max(n, len(maps))):
        roman = STAGE_ROMAN[i] if i < len(STAGE_ROMAN) else str(i + 1)
        name = maps[i] if i < len(maps) else "—"
        lines.append(f"**{roman}** · `{name}`")
    return "\n".join(lines)


def map_for_step(data: dict[str, Any], step_name: str) -> str | None:
    idx = STEP_MAP_INDEX.get(step_name)
    if idx is None:
        return None
    maps = get_kv_maps(data)
    if idx < 0 or idx >= len(maps):
        return None
    return maps[idx]


async def resolve_member(interaction: discord.Interaction) -> discord.Member | None:
    """Надёжно получить Member (с ролями) из slash-interaction."""
    if isinstance(interaction.user, discord.Member):
        return interaction.user
    guild = interaction.guild
    if guild is None:
        return None
    mid = interaction.user.id
    m = guild.get_member(mid)
    if m is not None:
        return m
    try:
        return await guild.fetch_member(mid)
    except Exception:
        return None


async def reply_interaction(
    interaction: discord.Interaction,
    *,
    embed: discord.Embed | None = None,
    content: str | None = None,
    ephemeral: bool = True,
) -> None:
    kwargs: dict[str, Any] = {"ephemeral": ephemeral}
    if embed is not None:
        kwargs["embed"] = embed
    if content is not None:
        kwargs["content"] = content
    try:
        if interaction.response.is_done():
            await interaction.followup.send(**kwargs)
        else:
            await interaction.response.send_message(**kwargs)
    except Exception as e:
        log(f"reply: {e}", "err")


def load_add_phrases() -> list[str]:
    """Строки из add_phrases.txt (пустые и #комменты пропускаем)."""
    if not ADD_PHRASES_FILE.exists():
        return []
    try:
        lines = ADD_PHRASES_FILE.read_text(encoding="utf-8").splitlines()
    except Exception as e:
        log(f"phrases: {e}", "err")
        return []
    out: list[str] = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        out.append(s)
    return out


async def send_random_add_phrase(channel: discord.abc.Messageable | None, delay: float = 5.0) -> None:
    """Через delay сек — случайная фраза из txt в этот чат."""
    if channel is None:
        return
    try:
        await asyncio.sleep(delay)
        phrases = load_add_phrases()
        if not phrases:
            return
        text = random.choice(phrases)
        if len(text) > 2000:
            text = text[:1997] + "…"
        await channel.send(text)
    except Exception:
        pass


def schedule_add_phrase(channel: discord.abc.Messageable | None) -> None:
    if channel is None:
        return
    asyncio.create_task(send_random_add_phrase(channel, 5.0))


async def do_roster_add(
    actor: discord.Member,
    target: discord.Member,
    game_nick: str,
) -> tuple[bool, discord.Embed]:
    """Общая логика /add и !add. (успех, embed)."""
    data_preview = load_db()
    if not can_use_roster(actor, data_preview):
        allowed = get_access_role_ids(data_preview)
        roles_txt = ", ".join(f"<@&{r}>" for r in allowed) if allowed else "_ролей нет — только админ_"
        return False, make_reply_embed(
            "❌  Нет доступа",
            f"Нужна роль: {roles_txt}\nили админ. `/access_add @роль`",
            color=COLOR_ERR,
        )

    game_nick = game_nick.strip()
    if not game_nick:
        return False, make_reply_embed("❌  Ошибка", "Укажи ник в **STALCRAFT**.", color=COLOR_ERR)
    if target.bot:
        return False, make_reply_embed("❌  Ошибка", "Ботов в состав добавлять нельзя.", color=COLOR_ERR)

    guild = actor.guild
    async with db_lock:
        data = load_db()
        for did, p in data.get("players", {}).items():
            if p.get("game_nick", "").lower() == game_nick.lower() and did != str(target.id):
                return False, make_reply_embed(
                    "❌  Ник уже занят",
                    f"Ник STALCRAFT `{game_nick}` уже привязан к <@{did}>.",
                    color=COLOR_ERR,
                )

        was = str(target.id) in data.get("players", {})
        prev = data.get("players", {}).get(str(target.id), {})
        old_came = prev.get("came", False)
        old_nick = prev.get("game_nick")
        data.setdefault("players", {})[str(target.id)] = {
            "discord_name": target.display_name,
            "discord_username": target.name,
            "game_nick": game_nick,
            "came": old_came,
            "in_voice": False,
            "squad": prev.get("squad"),
            "slot": prev.get("slot"),
        }
        # история грен едет с ником (без учёта регистра)
        if old_nick and old_nick != game_nick:
            _hist_move_ci(data.setdefault("grenade_history", {}), old_nick, game_nick)
        reindex_scan_histories(data)
        if guild and is_kv_live(data):
            voice_ids = collect_voice_member_ids(guild)
            if target.id in voice_ids:
                data["players"][str(target.id)]["in_voice"] = True
                if can_mark_came(data):
                    data["players"][str(target.id)]["came"] = True
        save_db(data)
        await upsert_status_messages(data)

    # tech = players.json (после unlock — Excel может подтормаживать)
    await sync_tech_sheet(data)

    action = "Обновлён" if was else "Добавлен"
    return True, make_reply_embed(
        f"✅  {action} в состав",
        (
            f"**Discord:** {target.mention}\n"
            f"**STALCRAFT:** `{game_nick}`\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"Игрок {'обновлён' if was else 'добавлен'} в базу."
        ),
        color=COLOR_OK,
    )


def find_player_id_by_game_nick(data: dict[str, Any], game_nick: str) -> str | None:
    """discord_id str по game_nick (без учёта регистра) или None."""
    want = (game_nick or "").strip().lower()
    if not want:
        return None
    for did, p in (data.get("players") or {}).items():
        gn = (p.get("game_nick") or "").strip().lower()
        if gn and gn == want:
            return str(did)
    return None


async def do_roster_remove(
    actor: discord.Member,
    target: discord.Member | None = None,
    game_nick: str | None = None,
) -> discord.Embed:
    """
    Удалить из состава:
      · по Discord (target)
      · или по нику STALCRAFT (game_nick) — как в Excel/таблице
    """
    data_preview = load_db()
    if not can_use_roster(actor, data_preview):
        allowed = get_access_role_ids(data_preview)
        roles_txt = ", ".join(f"<@&{r}>" for r in allowed) if allowed else "_ролей нет — только админ_"
        return make_reply_embed(
            "❌  Нет доступа",
            f"Нужна роль: {roles_txt}\nили админ. `/access_add @роль`",
            color=COLOR_ERR,
        )

    nick_arg = (game_nick or "").strip() or None
    if target is None and not nick_arg:
        return make_reply_embed(
            "❌  Ошибка",
            "Укажи **@Discord** или **ник STALCRAFT**.\n"
            "· `/remove @user` · `/remove ник:sosew`\n"
            "· `!remove @user` · `!remove sosew`",
            color=COLOR_ERR,
        )
    if target is not None and nick_arg:
        return make_reply_embed(
            "❌  Ошибка",
            "Укажи **либо** @Discord, **либо** ник — не оба сразу.",
            color=COLOR_ERR,
        )

    async with db_lock:
        data = load_db()
        players = data.get("players", {})

        if target is not None:
            key = str(target.id)
            if key not in players:
                return make_reply_embed(
                    "ℹ️  Нет в базе",
                    f"{target.mention} не найден в составе.",
                    color=COLOR_WAIT,
                )
        else:
            key = find_player_id_by_game_nick(data, nick_arg or "")
            if key is None:
                return make_reply_embed(
                    "ℹ️  Нет в базе",
                    f"Ник STALCRAFT `{nick_arg}` не найден в составе.\n"
                    f"Проверь написание или удали через `@Discord`.",
                    color=COLOR_WAIT,
                )

        removed = players.pop(key)
        gnick = removed.get("game_nick")
        if gnick:
            _hist_pop_ci(data.setdefault("grenade_history", {}), gnick)
        save_db(data)
        await upsert_status_messages(data)

    await sync_tech_sheet(data)

    ds_line = f"<@{key}>"
    return make_reply_embed(
        "🗑️  Удалён из состава",
        (
            f"**Discord:** {ds_line}\n"
            f"**Был ник DS:** `{removed.get('discord_name') or '—'}`\n"
            f"**STALCRAFT:** `{removed.get('game_nick') or nick_arg or '—'}`\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"Запись полностью убрана из базы."
        ),
        color=COLOR_ERR,
    )


# ---------------------------------------------------------------------------
# Stalcraft API (async aiohttp)
# ---------------------------------------------------------------------------
_token_cache: str | None = None
_token_expire_mono: float = 0.0
_http_session: aiohttp.ClientSession | None = None


async def get_http() -> aiohttp.ClientSession:
    global _http_session
    if _http_session is None or _http_session.closed:
        _http_session = aiohttp.ClientSession(
            timeout=aiohttp.ClientTimeout(total=25),
        )
    return _http_session


async def close_http() -> None:
    global _http_session
    if _http_session is not None and not _http_session.closed:
        await _http_session.close()
    _http_session = None


async def get_access_token() -> str | None:
    """OAuth app-token с кэшем."""
    global _token_cache, _token_expire_mono
    if _token_cache and time.monotonic() < _token_expire_mono:
        return _token_cache
    if not CLIENT_SECRET:
        log("нет STALCRAFT_CLIENT_SECRET", "err")
        return None
    url = "https://exbo.net/oauth/token"
    payload = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
    }
    try:
        session = await get_http()
        async with session.post(url, data=payload) as r:
            text = await r.text()
            if r.status == 200:
                data = await r.json(content_type=None)
                tok = data.get("access_token")
                exp = int(data.get("expires_in") or 3600)
                _token_cache = tok
                _token_expire_mono = time.monotonic() + max(60, min(exp - 60, 3600))
                return tok
            log(f"oauth HTTP {r.status}", "err")
    except Exception as e:
        log(f"oauth: {e}", "err")
    return None


def get_stat_value(stats_list: list, stat_id: str) -> int:
    for stat in stats_list:
        if stat.get("id") == stat_id:
            try:
                return int(float(stat.get("value") or 0))
            except (TypeError, ValueError):
                return 0
    return 0


# eAPI: gre-thr — брошенные гранаты (lifetime)
STAT_GRE = "gre-thr"


async def fetch_player_grenades(
    session: aiohttp.ClientSession,
    token: str,
    nickname: str,
) -> tuple[str, int | None]:
    """Один запрос profile → gre-thr."""
    safe = quote(str(nickname), safe="")
    url = f"https://eapi.stalcraft.net/{REGION}/character/by-name/{safe}/profile"
    headers = {"Authorization": f"Bearer {token}", "Cache-Control": "no-cache"}
    try:
        async with session.get(url, headers=headers) as r:
            if r.status == 200:
                body = await r.json(content_type=None)
                stats = body.get("stats", [])
                return nickname, get_stat_value(stats, STAT_GRE)
            if r.status != 404:
                log(f"API {r.status} · {nickname}", "warn")
    except Exception as e:
        log(f"API {nickname}: {e}", "err")
    return nickname, None


async def scan_grenades(nicks: list[str]) -> dict[str, int | None]:
    """Параллельный скан gre-thr (до 5 одновременно)."""
    if not nicks:
        return {}
    token = await get_access_token()
    if not token:
        log("нет токена API — скан отменён", "err")
        return {n: None for n in nicks}

    session = await get_http()
    sem = asyncio.Semaphore(5)

    async def one(nick: str) -> tuple[str, int | None]:
        async with sem:
            await asyncio.sleep(0.08)
            return await fetch_player_grenades(session, token, nick)

    results = await asyncio.gather(*[one(n) for n in nicks])
    return {name: gre for name, gre in results}


# ---------------------------------------------------------------------------
# Excel: отряды (squads.xlsx / SHEET_PATH)
# ---------------------------------------------------------------------------
_sheet_last_mtime: float | None = None
_sheet_last_sig: tuple | None = None
_sheet_last_error: str | None = None


def _sheet_lock_path(path: Path | None = None) -> Path:
    """Excel lock-файл ~$name.xlsx — значит таблица открыта в Excel."""
    path = path or SHEET_PATH
    return path.with_name(f"~${path.name}")


def _entries_signature(entries: list[dict[str, Any]]) -> tuple:
    return tuple(
        (
            (e.get("game_nick") or "").lower(),
            e.get("squad"),
            e.get("slot"),
            e.get("discord_id") or "",
        )
        for e in sorted(
            entries,
            key=lambda x: (
                x.get("squad") is None,
                x.get("squad") or 0,
                x.get("slot") or 0,
                (x.get("game_nick") or "").lower(),
            ),
        )
    )


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _parse_discord_id(val: Any) -> str | None:
    s = _cell_str(val)
    if not s:
        return None
    s = s.replace(" ", "")
    m = re.search(r"\d{15,22}", s)
    if not m:
        return None
    return m.group(0)


def _parse_int_cell(val: Any) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        try:
            return int(val)
        except (TypeError, ValueError):
            return None
    s = _cell_str(val)
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _is_subs_header(val: Any) -> bool:
    s = _cell_str(val).lower().replace("ё", "е")
    return s in ("замены", "замена", "subs", "sub", "reserve", "резерв", "скам", "чемпионы")


def _squad_header_id(val: Any) -> int | None:
    """Если ячейка — заголовок отряда (1..20 или Замены), вернуть id; иначе None."""
    if _is_subs_header(val):
        return 99
    n = _parse_int_cell(val)
    # только короткие номера отрядов, не ники-числа
    if n is not None and 1 <= n <= 20 and len(_cell_str(val)) <= 3:
        return n
    return None


_TECH_SHEET_NAMES = ("tech", "ids", "id", "база", "discord", "mapping", "привязки", "nicks")


def _find_tech_sheet(wb):
    """Найти лист tech / ids / … или None."""
    for name in wb.sheetnames:
        if name.strip().lower() in _TECH_SHEET_NAMES:
            return wb[name]
    return None


def _read_tech_nick_map(wb) -> dict[str, str]:
    """
    Лист tech / ids / база / discord:
    колонки game_nick | discord_id  (порядок любой по заголовку).
    """
    sheet = _find_tech_sheet(wb)
    if sheet is None:
        return {}

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [_cell_str(x).lower() for x in (rows[0] or ())]
    has_header = any(
        h in ("game_nick", "nick", "ник", "stalcraft", "discord_id", "discord", "id")
        for h in header
    )
    if has_header:
        c_nick = next(
            (i for i, h in enumerate(header) if h in ("game_nick", "nick", "ник", "stalcraft", "name")),
            0,
        )
        c_did = next(
            (i for i, h in enumerate(header) if h in ("discord_id", "discord", "id", "ds")),
            1,
        )
        data_rows = rows[1:]
    else:
        c_nick, c_did = 0, 1
        data_rows = rows

    out: dict[str, str] = {}
    for row in data_rows:
        if not row:
            continue
        nick = _cell_str(row[c_nick] if c_nick < len(row) else "")
        did = _parse_discord_id(row[c_did] if c_did < len(row) else None)
        # допускаем и обратный порядок, если id в первой колонке
        if not did and not has_header:
            did = _parse_discord_id(row[0] if row else None)
            nick = _cell_str(row[1] if len(row) > 1 else "")
        if nick and did:
            out[nick.lower()] = did
    return out


def write_tech_sheet_from_players(
    data: dict[str, Any] | None = None,
    path: Path | None = None,
) -> tuple[bool, str]:
    """
    Полностью перезаписывает лист tech из players.
    Колонки: game_nick | discord_id (id всегда текстом, без float).
    Лист1 (красивая таблица) не трогает.
    """
    path = path or SHEET_PATH
    if data is None:
        data = load_db()
    if not path.exists():
        return False, f"tech: нет файла {path.name}"

    rows: list[tuple[str, str]] = []
    for did, p in (data.get("players") or {}).items():
        nick = (p.get("game_nick") or "").strip()
        did_s = str(did).strip()
        if not nick or not did_s.isdigit():
            continue
        rows.append((nick, did_s))
    rows.sort(key=lambda x: x[0].lower())

    try:
        wb = load_workbook(path)
        sheet = _find_tech_sheet(wb)
        if sheet is None:
            sheet = wb.create_sheet("tech")

        # очистить старые строки (заголовок + данные)
        if sheet.max_row and sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row)

        sheet.cell(row=1, column=1, value="game_nick")
        sheet.cell(row=1, column=2, value="discord_id")
        for i, (nick, did_s) in enumerate(rows, start=2):
            sheet.cell(row=i, column=1, value=nick)
            id_cell = sheet.cell(row=i, column=2, value=did_s)
            id_cell.number_format = "@"  # текст — Discord id не в float

        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 24

        wb.save(path)
        wb.close()
    except PermissionError:
        return False, f"tech: файл занят Excel — закрой или Ctrl+S ({path.name})"
    except Exception as e:
        return False, f"tech: ошибка записи: {e}"

    return True, f"tech: {len(rows)} ников → {path.name}"


_tech_pending: bool = False
_tech_lock_warned: bool = False


async def sync_tech_sheet(data: dict[str, Any] | None = None) -> None:
    """
    Обновить лист tech. Если Excel держит файл — ставим pending и
    пробуем снова на следующем sheet_sync (без спама в лог).
    """
    global _tech_pending, _tech_lock_warned
    loop = asyncio.get_running_loop()
    ok, msg = await loop.run_in_executor(None, write_tech_sheet_from_players, data)
    if ok:
        _tech_pending = False
        _tech_lock_warned = False
        log(msg, "ok")
        return
    # файл занят — отложим
    if "занят" in msg.lower() or "permission" in msg.lower():
        _tech_pending = True
        if not _tech_lock_warned:
            _tech_lock_warned = True
            log("tech: Excel открыт — tech допишется после Ctrl+S / закрытия", "warn")
        return
    log(msg, "warn")


def _read_roster_table(ws) -> list[dict[str, Any]]:
    """Классика: squad | slot | discord_id | game_nick."""
    rows_iter = ws.iter_rows(values_only=True)
    header_raw = next(rows_iter, None)
    if not header_raw:
        return []
    header = [_cell_str(h).lower() for h in header_raw]

    def col(*names: str, default: int | None = None) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        return default

    # без discord_id в заголовке — это не roster-формат
    if col("discord_id", "discord", "id", "discordid") is None and "discord_id" not in header:
        # если есть game_nick+squad — можно без id (резолв по нику)
        if col("game_nick", "nick", "ник", "stalcraft") is None:
            return []

    c_squad = col("squad", "отряд", "otr", default=0)
    c_slot = col("slot", "место", "слот", default=1)
    c_did = col("discord_id", "discord", "id", "discordid", default=2)
    c_nick = col("game_nick", "nick", "ник", "stalcraft", "game", default=3)
    c_dname = col("discord_name", "name", "имя", default=4)

    entries: list[dict[str, Any]] = []
    for row in rows_iter:
        if not row:
            continue

        def get(i: int | None) -> Any:
            if i is None or i >= len(row):
                return None
            return row[i]

        did = _parse_discord_id(get(c_did)) if c_did is not None else None
        nick = _cell_str(get(c_nick))
        squad = _parse_int_cell(get(c_squad))
        slot = _parse_int_cell(get(c_slot))
        dname = _cell_str(get(c_dname)) if c_dname is not None else ""
        if not did and not nick:
            continue
        entries.append(
            {
                "discord_id": did,
                "game_nick": nick,
                "squad": squad,
                "slot": slot,
                "discord_name": dname,
            }
        )
    return entries


# Жёсткая карта ячеек ДИТЯ22 (openpyxl: row/col 1-based, C=3 … F=6)
#   C6:C10 → отряд 1
#   D6:D10 → отряд 2
#   E6:E10 → отряд 3
#   C12:C16 → отряд 4
#   D12:D16 → отряд 5
#   E12:E16 → отряд 6
#   F6:F16  → Чемпионы (замены)
FIXED_SQUAD_RANGES: list[tuple[int, int, int, int, int]] = [
    # squad_id, col, row_from, row_to
    (1, 3, 6, 10),
    (2, 4, 6, 10),
    (3, 5, 6, 10),
    (4, 3, 12, 16),
    (5, 4, 12, 16),
    (6, 5, 12, 16),
    (99, 6, 6, 16),  # F6:F16 Чемпионы
]


def _read_pretty_grid(ws) -> tuple[list[dict[str, Any]], dict[str, str]]:
    """
    Читает только закреплённые диапазоны ячеек (см. FIXED_SQUAD_RANGES).
    Заметки вне этих ячеек (КТРЛ+S и т.п.) игнорируются.
    """
    entries: list[dict[str, Any]] = []
    squad_names: dict[str, str] = {
        "1": "Отряд 1",
        "2": "Отряд 2",
        "3": "Отряд 3",
        "4": "Отряд 4",
        "5": "Отряд 5",
        "6": "Отряд 6",
        "99": "Чемпионы",
    }

    for sid, col, r0, r1 in FIXED_SQUAD_RANGES:
        slot = 0
        for row in range(r0, r1 + 1):
            raw = ws.cell(row=row, column=col).value
            nick = _cell_str(raw)
            if not nick:
                continue
            # пропуск случайных номеров-заголовков в ячейке
            if _squad_header_id(nick) is not None:
                continue
            if _is_subs_header(nick):
                continue
            slot += 1
            entries.append(
                {
                    "discord_id": None,
                    "game_nick": nick,
                    "squad": sid,
                    "slot": slot,
                    "discord_name": "",
                }
            )

    return entries, squad_names


def read_squad_sheet(
    path: Path | None = None,
) -> tuple[list[dict[str, Any]], dict[str, str], dict[str, str]]:
    """
    Читает Excel.
    Returns: (entries, squad_names, tech_map nick_lower→discord_id)

    Поддерживает:
    1) roster: squad, slot, discord_id?, game_nick
    2) красивая сетка: столбцы 1 2 3 Замены / 4 5 6 + ники
    3) лист tech: ник ↔ discord_id
    """
    path = path or SHEET_PATH
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    # read_only=False — надёжнее для маленьких файлов / нескольких листов
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        tech_map = _read_tech_nick_map(wb)

        squad_names: dict[str, str] = {}
        if "squads" in wb.sheetnames:
            ws_n = wb["squads"]
            rows_list = list(ws_n.iter_rows(values_only=True))
            if rows_list:
                header_n = [_cell_str(x).lower() for x in (rows_list[0] or ())]
                has_header = any(h in ("squad", "id", "отряд", "name", "имя") for h in header_n)
                data_rows = rows_list[1:] if has_header else rows_list
                if has_header:
                    idx_id = next((i for i, h in enumerate(header_n) if h in ("squad", "id", "отряд")), 0)
                    idx_name = next((i for i, h in enumerate(header_n) if h in ("name", "имя", "title")), 1)
                else:
                    idx_id, idx_name = 0, 1
                for row in data_rows:
                    if not row:
                        continue
                    sid = _parse_int_cell(row[idx_id] if idx_id < len(row) else None)
                    if sid is None:
                        continue
                    name = _cell_str(row[idx_name] if idx_name < len(row) else "") or f"Отряд {sid}"
                    squad_names[str(sid)] = name

        # основной лист: roster или первый / Лист1
        if "roster" in wb.sheetnames:
            main = wb["roster"]
        elif "Лист1" in wb.sheetnames:
            main = wb["Лист1"]
        else:
            main = wb[wb.sheetnames[0]]

        # пробуем табличный roster
        entries = _read_roster_table(main)
        # если пусто или нет ни одной строки с ником — красивая сетка
        if not entries:
            grid_entries, grid_names = _read_pretty_grid(main)
            entries = grid_entries
            for k, v in grid_names.items():
                squad_names.setdefault(k, v)
        else:
            # roster мог быть ложно распознан с пустыми id — если все без nick+squad grid
            if entries and all(not e.get("discord_id") for e in entries):
                # возможно это не roster; проверим сетку
                grid_entries, grid_names = _read_pretty_grid(main)
                if len(grid_entries) > len(entries):
                    entries = grid_entries
                    for k, v in grid_names.items():
                        squad_names.setdefault(k, v)

        return entries, squad_names, tech_map
    finally:
        wb.close()


def _hist_get(hist: dict[str, Any], nick: str) -> Any | None:
    """Достать запись history по нику (без учёта регистра)."""
    if not hist or not nick:
        return None
    if nick in hist:
        return hist[nick]
    nl = nick.lower()
    for k, v in hist.items():
        if str(k).lower() == nl:
            return v
    return None


def _hist_pop_ci(hist: dict[str, Any], nick: str) -> Any | None:
    if not hist or not nick:
        return None
    if nick in hist:
        return hist.pop(nick)
    nl = nick.lower()
    for k in list(hist.keys()):
        if str(k).lower() == nl:
            return hist.pop(k)
    return None


def _hist_move_ci(hist: dict[str, Any], old_nick: str, new_nick: str) -> bool:
    """Перенести историю грен со старого ника на новый (регистр не важен)."""
    if not old_nick or not new_nick:
        return False
    if old_nick.lower() == new_nick.lower():
        # только выровнять ключ под актуальный регистр
        val = _hist_pop_ci(hist, old_nick)
        if val is not None:
            hist[new_nick] = val
            return True
        return False
    val = _hist_pop_ci(hist, old_nick)
    if val is None:
        return False
    # если на новом нике уже есть данные — не затираем
    if _hist_get(hist, new_nick) is None:
        hist[new_nick] = val
    return True


def reindex_scan_histories(data: dict[str, Any]) -> list[str]:
    """
    Ключи grenade_history → актуальные game_nick из players.
    Чтобы при смене регистра / правке Excel данные не «терялись».
    """
    changes: list[str] = []
    players = data.get("players") or {}
    preferred: dict[str, str] = {}
    for p in players.values():
        gn = (p.get("game_nick") or "").strip()
        if gn:
            preferred[gn.lower()] = gn

    # выкинуть устаревший combat_history (HS% больше не используем)
    if "combat_history" in data:
        del data["combat_history"]

    hist = data.get("grenade_history")
    if not isinstance(hist, dict) or not hist:
        return changes
    new_h: dict[str, Any] = {}
    for old_key, val in hist.items():
        if not old_key:
            continue
        pref = preferred.get(str(old_key).lower())
        target = pref or str(old_key)
        if target in new_h and pref and str(old_key) != target:
            changes.append(f"hist merge `{old_key}` → `{target}`")
            continue
        if pref and pref != old_key:
            changes.append(f"hist `{old_key}` → `{pref}`")
        new_h[target] = val
    data["grenade_history"] = new_h
    return changes


def apply_sheet_entries(
    data: dict[str, Any],
    entries: list[dict[str, Any]],
    squad_names: dict[str, str],
    tech_map: dict[str, str] | None = None,
) -> tuple[dict[str, Any], list[str]]:
    """
    Применяет Excel к players.
    Discord резолвится: entry.discord_id → tech_map[ник] → players по game_nick.
    Красивая таблица может содержать только ники.
    Не удаляет игроков из базы.
    Грены привязаны к нику: при смене ника/перестановке в Excel едут с игроком.
    """
    changes: list[str] = []
    players = data.setdefault("players", {})
    history = data.setdefault("grenade_history", {})
    tech_map = tech_map or {}

    if squad_names:
        data["squad_names"] = {str(k): str(v) for k, v in squad_names.items()}
    else:
        data.setdefault("squad_names", {})

    # ник (lower) → discord_id из базы
    nick_to_id: dict[str, str] = {}
    for did, p in players.items():
        gn = (p.get("game_nick") or "").strip()
        if gn:
            nick_to_id[gn.lower()] = did

    # тех-лист перекрывает / дополняет
    for nick_l, did in tech_map.items():
        nick_to_id[nick_l] = did
        if did not in players:
            # создать запись из tech (ник = ключ как в таблице — восстановим регистр из entries позже)
            players[did] = {
                "discord_name": "",
                "discord_username": "",
                "game_nick": nick_l,  # временный; поправим ниже
                "came": False,
                "in_voice": False,
                "squad": None,
                "slot": None,
            }
            changes.append(f"+ tech <@{did}> `{nick_l}`")

    seen_ids: set[str] = set()
    nick_owner: dict[str, str] = {}

    for e in entries:
        nick = (e.get("game_nick") or "").strip()
        did = e.get("discord_id") or None
        if isinstance(did, str):
            did = did.strip() or None
        squad = e.get("squad")
        slot = e.get("slot")
        dname = (e.get("discord_name") or "").strip()

        # резолв Discord по нику
        if not did and nick:
            did = nick_to_id.get(nick.lower())
        if not did:
            if nick:
                changes.append(f"⚠️ ник `{nick}` нет в базе — `/add` или лист tech")
            continue

        if nick:
            other = nick_owner.get(nick.lower())
            if other and other != did:
                changes.append(f"⚠️ ник `{nick}` дважды — пропуск")
                continue
            nick_owner[nick.lower()] = did

        seen_ids.add(did)
        p = players.get(did)
        if p is None:
            if not nick:
                changes.append(f"⚠️ id `{did}` без ника — пропуск")
                continue
            players[did] = {
                "discord_name": dname or nick,
                "discord_username": "",
                "game_nick": nick,
                "came": False,
                "in_voice": False,
                "squad": squad,
                "slot": slot,
            }
            nick_to_id[nick.lower()] = did
            changes.append(f"+ <@{did}> `{nick}` отр.{squad}/{slot}")
            continue

        old_nick = (p.get("game_nick") or "").strip()
        old_squad = p.get("squad")
        old_slot = p.get("slot")

        # ник из красивой таблицы = актуальный game_nick
        if nick and nick != old_nick:
            conflict = False
            for oid, op in players.items():
                if oid != did and (op.get("game_nick") or "").lower() == nick.lower():
                    changes.append(f"⚠️ `{nick}` уже у <@{oid}> — ник <@{did}> не сменён")
                    conflict = True
                    break
            if not conflict:
                if old_nick:
                    _hist_move_ci(history, old_nick, nick)
                    nick_to_id.pop(old_nick.lower(), None)
                p["game_nick"] = nick
                nick_to_id[nick.lower()] = did
                changes.append(f"ник <@{did}>: `{old_nick}` → `{nick}`")

        if dname and dname != p.get("discord_name"):
            p["discord_name"] = dname

        if squad != old_squad or slot != old_slot:
            p["squad"] = squad
            p["slot"] = slot
            changes.append(
                f"отряд `{p.get('game_nick') or did}`: {old_squad}/{old_slot} → {squad}/{slot}"
            )

    for did, p in players.items():
        if did in seen_ids:
            continue
        if p.get("squad") is not None or p.get("slot") is not None:
            p["squad"] = None
            p["slot"] = None
            changes.append(f"сброс отряда `{p.get('game_nick') or did}` (нет в таблице)")

    # выровнять ключи грен под актуальные ники (после перестановок в Excel)
    for msg in reindex_scan_histories(data):
        changes.append(msg)

    return data, changes


def squad_label(data: dict[str, Any], squad: Any) -> str:
    """Коротко: Отряд 1: … Отряд 6:, Чемпионы:, Без отряда:"""
    if squad is None:
        return "Без отряда:"
    try:
        sid = int(squad)
    except (TypeError, ValueError):
        sid = squad
    if sid == 99 or str(squad) == "99":
        return "Чемпионы:"
    names = data.get("squad_names") or {}
    name = (names.get(str(squad)) or "").strip()
    # не дублировать «Отряд 1 · Отряд 1»
    if not name or name == f"Отряд {sid}" or name.lower() in ("замены", "замена"):
        return f"Отряд {sid}:"
    if name.lower() in ("чемпионы", "champions"):
        return "Чемпионы:"
    return f"Отряд {sid} · {name}:"


def player_squad_sort_key(item: tuple[str, dict[str, Any]]) -> tuple:
    _did, p = item
    sq = p.get("squad")
    sl = p.get("slot")
    nick = (p.get("game_nick") or "").lower()
    has = 0 if sq is not None else 1
    try:
        sq_n = int(sq) if sq is not None else 999
    except (TypeError, ValueError):
        sq_n = 999
    try:
        sl_n = int(sl) if sl is not None else 999
    except (TypeError, ValueError):
        sl_n = 999
    return (has, sq_n, sl_n, nick)


# ---------------------------------------------------------------------------
# Discord bot
# ---------------------------------------------------------------------------
intents = discord.Intents.default()
intents.members = True
intents.voice_states = True
intents.guilds = True
# для запасных !add / !remove (если slash скрыт правами Discord)
intents.message_content = True

bot = commands.Bot(
    command_prefix=commands.when_mentioned_or("!"),
    intents=intents,
    help_command=None,
)


# ---------------------------------------------------------------------------
# Вспомогательное оформление
# ---------------------------------------------------------------------------
def now_msk() -> datetime:
    return datetime.now(MSK)


def today_msk_str() -> str:
    return now_msk().strftime("%Y-%m-%d")


def kv_dt_from_session(data: dict[str, Any] | None = None) -> datetime:
    """Дата сессии КВ → datetime (для выбора вс / чт–сб расписания)."""
    raw = None
    if data:
        raw = data.get("grenade_date") or data.get("session_date")
    if raw:
        try:
            y, m, d = map(int, str(raw).split("-")[:3])
            return datetime(y, m, d, tzinfo=MSK)
        except Exception:
            pass
    return now_msk()


def is_cw_day(dt: datetime | None = None) -> bool:
    """КВ-дни: чт / пт / сб / вс."""
    return (dt or now_msk()).weekday() in CW_WEEKDAYS


def is_kv_finished(data: dict[str, Any]) -> bool:
    if data.get("kv_finished"):
        return True
    last = data.get("last_grenade_step")
    if not last:
        return False
    return last == kv_final_step_id(kv_dt_from_session(data))


def is_kv_live(data: dict[str, Any] | None = None) -> bool:
    """
    Живое окно КВ для отображения явки (не 💤):
    день КВ, сессия открыта, с явки до финала.
    """
    if data is None:
        data = load_db()
    n = now_msk()
    if not is_cw_day(n):
        return False
    if n.time() < kv_start(n):
        return False
    if not data.get("kv_session_active"):
        return False
    if is_kv_finished(data):
        return False
    return True


def can_mark_came(data: dict[str, Any] | None = None) -> bool:
    """Можно ставить ✅: только в окне явки (до старта 1 этапа)."""
    if data is None:
        data = load_db()
    if not is_kv_live(data):
        return False
    return now_msk().time() <= kv_attendance_end()


def format_kv_date_label(data: dict[str, Any]) -> str:
    """четверг 18.07"""
    from datetime import date as date_cls

    raw = data.get("grenade_date") or data.get("session_date")
    if raw:
        try:
            y, m, d = map(int, str(raw).split("-")[:3])
            dt = date_cls(y, m, d)
            return f"{WEEKDAYS_RU[dt.weekday()]} {d:02d}.{m:02d}"
        except Exception:
            pass
    n = now_msk()
    return f"{WEEKDAYS_RU[n.weekday()]} {n.day:02d}.{n.month:02d}"


def _minutes_of(t: dt_time) -> int:
    return int(t.hour) * 60 + int(t.minute)


def _minutes_late(now_t: dt_time, target: dt_time) -> int:
    """Сколько минут сейчас после target (отрицательно = ещё рано)."""
    return _minutes_of(now_t) - _minutes_of(target)


def pad_nick(nick: str, width: int = 14) -> str:
    """Подгонка ника под monospace."""
    nick = (nick or "").replace("\n", " ").strip()
    if len(nick) > width:
        return nick[: max(1, width - 1)] + "…"
    return nick.ljust(width)


def cell(val: str | int | None, width: int = 4) -> str:
    if val is None or val == "—":
        s = "-"
    else:
        s = str(val)
    return s.rjust(width)


def split_field_chunks(lines: list[str], limit: int = 900) -> list[str]:
    """
    Режет строки на куски ≤ limit.
    Заголовок «Отряд N:» не отрывается от игроков: новый кусок лучше начинать с отряда.
    """
    chunks: list[str] = []
    cur: list[str] = []

    def cur_len() -> int:
        if not cur:
            return 0
        return sum(len(x) for x in cur) + max(0, len(cur) - 1)

    def flush() -> None:
        if cur:
            chunks.append("\n".join(cur))
            cur.clear()

    for line in lines:
        if not line and not cur:
            continue
        extra = len(line) + (1 if cur else 0)
        # если не влезает — сброс (кроме пустого cur)
        if cur and cur_len() + extra > limit:
            flush()
        # если строка-отряд и кусок уже большой (>60%) — лучше новый field
        is_squad = bool(line) and line.endswith(":") and not line.startswith("ник")
        if cur and is_squad and cur_len() > limit * 0.55:
            flush()
        cur.append(line)

    flush()
    return chunks or ["-"]


def make_reply_embed(
    title: str,
    description: str,
    color: int = COLOR_INFO,
    footer: str | None = None,
) -> discord.Embed:
    emb = discord.Embed(title=title, description=description, color=color, timestamp=now_msk())
    emb.set_footer(text=footer or "КВ STALCRAFT · время МСК")
    return emb


# ---------------------------------------------------------------------------
# Основные embeds: ОНЛАЙН / ГРАНАТЫ
# ---------------------------------------------------------------------------
def format_online_embed(data: dict[str, Any]) -> discord.Embed:
    players = data.get("players", {})
    live = is_kv_live(data)
    sleep = not live
    n = now_msk()
    start_s = kv_start(n).strftime("%H:%M")
    att_end_s = kv_attendance_end(n).strftime("%H:%M")

    # статус зависит от дня/времени МСК и флагов сессии
    if live:
        if can_mark_came(data):
            status = f"🟢 Явка · до {att_end_s} МСК"
        else:
            status = "🟢 КВ идёт · грены по этапам"
        color = COLOR_ONLINE
    elif is_kv_finished(data) and data.get("session_date") == today_msk_str():
        status = "🏁 КВ сегодня завершён"
        color = COLOR_WAIT
    elif is_cw_day(n) and n.time() < kv_start(n):
        status = f"⏳ Ожидание · явка с {start_s} МСК"
        color = COLOR_WAIT
    elif is_cw_day(n) and n.time() >= kv_start(n) and not data.get("kv_session_active"):
        status = f"⏳ День КВ · сессия не открыта (жди тик или `/reset_session`)"
        color = COLOR_WAIT
    else:
        status = f"⏳ Вне КВ · чт–сб с 19:30 · вс с 18:30 МСК"
        color = COLOR_WAIT

    head = (
        f"{status}\n"
        f"\n"
        f"{SEP_LINE}\n"
        f"✅ пришёл  ·  ❌ не был\n"
        f"🔊 в войсе  ·  🔇 не в войсе\n"
        f"💤 сон      ·  вне КВ\n"
        f"{SEP_LINE}\n"
        f"\n"
    )

    embed = discord.Embed(
        title="📡  ЯВКА НА КВ",
        description=head,
        color=color,
    )

    if not players:
        embed.description = head + "_пусто · `/add`_"
        return embed

    # только валидные ники
    items = [
        (did, p)
        for did, p in players.items()
        if (p.get("game_nick") or "").strip()
    ]
    items = sorted(
        items,
        key=lambda x: (
            player_squad_sort_key(x)[0],
            player_squad_sort_key(x)[1],
            player_squad_sort_key(x)[2],
            0 if x[1].get("came") else 1,
            0 if x[1].get("in_voice") else 1,
            (x[1].get("game_nick") or "").lower(),
        ),
    )

    # Группы: (squad_key, title) → lines
    groups: list[tuple[Any, str, list[str]]] = []
    cur_key: Any = object()
    cur_title = ""
    cur_lines: list[str] = []

    def flush() -> None:
        nonlocal cur_lines
        if cur_lines:
            groups.append((cur_key, cur_title, cur_lines))
            cur_lines = []

    for did, p in items:
        gnick = (p.get("game_nick") or "").strip().replace("`", "'").replace("|", "/")
        if not gnick:
            continue
        sq = p.get("squad")
        sq_key = sq if sq is not None else None
        if sq_key != cur_key:
            flush()
            cur_key = sq_key
            cur_title = squad_label(data, sq)

        try:
            sl = int(p["slot"]) if p.get("slot") is not None else None
        except (TypeError, ValueError):
            sl = None
        slot_txt = f"{sl}. " if sl is not None else ""
        mention = f"<@{did}>"

        if sleep:
            cur_lines.append(f"💤  {slot_txt}`{gnick}` | {mention}")
        else:
            came = bool(p.get("came"))
            inv = bool(p.get("in_voice"))
            mark = "✅" if came else "❌"
            voice = "🔊" if inv else "🔇"
            cur_lines.append(f"{mark} {voice}  {slot_txt}`{gnick}` | {mention}")
    flush()

    # description: шапка + сколько влезет; остальное — field по отрядам
    body_parts: list[str] = []
    field_groups: list[tuple[str, list[str]]] = []
    budget = 3800 - len(head)

    for _key, title, glines in groups:
        block = f"**{title}**\n" + "\n".join(glines)
        # пустая строка между отрядами (в description)
        trial = ("\n\n".join(body_parts + [block])) if body_parts else block
        if not field_groups and len(trial) <= budget:
            body_parts.append(block)
        else:
            field_groups.append((title, glines))

    embed.description = head + ("\n\n".join(body_parts) if body_parts else "")
    for gi, (title, glines) in enumerate(field_groups):
        # в fields: отступ сверху через пустую строку в value (кроме самого первого field,
        # если description уже закончилась на отряде)
        chunks = split_field_chunks(glines, limit=1000)
        for i, chunk in enumerate(chunks):
            name = title if i == 0 else f"{title} · {i + 1}"
            if len(name) > 256:
                name = name[:253] + "…"
            val = chunk or "—"
            if i == 0 and (body_parts or gi > 0):
                # визуальный зазор: Discord fields и так разделены; name = заголовок отряда
                pass
            embed.add_field(name=name, value=val, inline=False)
    return embed


def _grenade_row_stats(
    history: dict, gnick: str, *, four_stages: bool = False
) -> tuple[int | None, int | None, int | None, int | None, int | None]:
    """I, II, III, [IV], ИТОГ."""
    s = _hist_get(history, gnick) or {}
    if not isinstance(s, dict):
        s = {}
    v0 = s.get("20:00")
    v1 = s.get("20:25")
    v2 = s.get("20:50")
    v3 = s.get("21:20")
    v4 = s.get("21:40")

    def stage_diff(a, b) -> int | None:
        if a is None or b is None:
            return None
        d = int(a) - int(b)
        return d if d >= 0 else None

    e1 = stage_diff(v1, v0)
    e2 = stage_diff(v2, v1) if v1 is not None else stage_diff(v2, v0)
    e3 = (
        stage_diff(v3, v2)
        if v2 is not None
        else (stage_diff(v3, v1) if v1 is not None else stage_diff(v3, v0))
    )
    e4 = None
    if four_stages:
        if v3 is not None:
            e4 = stage_diff(v4, v3)
        elif v2 is not None:
            e4 = stage_diff(v4, v2)
        else:
            e4 = stage_diff(v4, v0)

    parts = [x for x in (e1, e2, e3, e4) if x is not None]
    if parts:
        total: int | None = sum(parts)
    elif v0 is not None:
        total = 0
    else:
        total = None
    return e1, e2, e3, e4, total


def format_grenades_embed(data: dict[str, Any]) -> discord.Embed:
    """
    Классическое оформление:
      **Отряд 1:**
      `ник          I      II     III    ИТОГ`
    Только дельты gre-thr (без HS%).
    """
    players = data.get("players", {})
    history = data.get("grenade_history", {}) or {}
    date_label = format_kv_date_label(data)
    four = _weekday(kv_dt_from_session(data)) == 6

    embed = discord.Embed(
        title="💣  ГРАНАТЫ",
        color=COLOR_GRENADES,
    )

    nick_to_player: dict[str, tuple[str, dict[str, Any]]] = {}
    for did, p in players.items():
        gn = (p.get("game_nick") or "").strip()
        if gn:
            nick_to_player[gn] = (did, p)
            nick_to_player[gn.lower()] = (did, p)

    # Всегда состав отрядов 1–6 + Чемпионы (как в Excel); «Без отряда» не показываем.
    # Цифры висят после КВ до следующего 19:30 (обнуление в ensure_session_reset).
    nick_set: dict[str, str] = {}  # lower → display
    for p in players.values():
        if not is_scan_roster_player(p):
            continue
        gn = (p.get("game_nick") or "").strip()
        if gn:
            nick_set[gn.lower()] = gn
    # сироты history (ник был в отряде, потом убрали) — не тащим

    nick_list = list(nick_set.values())
    maps = get_kv_maps(data)
    if not nick_list:
        embed.description = "_чисто · нет ников в отрядах 1–6 / Чемпионы_"
        embed.set_footer(text=date_label)
        return embed

    def nick_sort_key(gnick: str) -> tuple:
        pair = nick_to_player.get(gnick) or nick_to_player.get(gnick.lower())
        if pair:
            return player_squad_sort_key(pair)
        return (1, 999, 999, gnick.lower())

    nick_list = sorted(nick_list, key=nick_sort_key)

    nick_w = 14
    cw = 5
    col_gap = "  "  # между I / II / III / IV
    tot_gap = "     "  # перед ИТОГ побольше

    def _cols(*cells: str, with_total: str | None = None) -> str:
        mid = col_gap.join(cells)
        if with_total is None:
            return mid
        return mid + tot_gap + with_total

    h_i = f"{'I':>{cw}}"
    h_ii = f"{'II':>{cw}}"
    h_iii = f"{'III':>{cw}}"
    h_iv = f"{'IV':>{cw}}"
    h_tot = f"{'ИТОГ':>{cw}}"
    dash = "─" * cw

    if four:
        hdr = f"`{'Ник':<{nick_w}} {_cols(h_i, h_ii, h_iii, h_iv, with_total=h_tot)}`"
        sep = f"`{'─' * nick_w} {_cols(dash, dash, dash, dash, with_total=dash)}`"
    else:
        hdr = f"`{'Ник':<{nick_w}} {_cols(h_i, h_ii, h_iii, with_total=h_tot)}`"
        sep = f"`{'─' * nick_w} {_cols(dash, dash, dash, with_total=dash)}`"

    body: list[str] = [hdr, sep]
    last_sq: Any = object()
    for gnick in nick_list:
        pair = nick_to_player.get(gnick) or nick_to_player.get(gnick.lower())
        sq = pair[1].get("squad") if pair else None
        if sq != last_sq:
            # пустая строка между отрядами (не перед первым)
            if last_sq is not object():
                body.append("")
            last_sq = sq
            body.append(f"**{squad_label(data, sq)}**")

        e1, e2, e3, e4, total = _grenade_row_stats(history, gnick, four_stages=four)
        nick_show = pad_nick(gnick, nick_w)
        if four:
            row = (
                f"`{nick_show} "
                f"{_cols(cell(e1, cw), cell(e2, cw), cell(e3, cw), cell(e4, cw), with_total=cell(total, cw))}`"
            )
        else:
            row = (
                f"`{nick_show} "
                f"{_cols(cell(e1, cw), cell(e2, cw), cell(e3, cw), with_total=cell(total, cw))}`"
            )
        body.append(row)

    # Одна description ≤4096 — без «·2/·3» и разрывов посреди отряда.
    # Если не влезло — режем только по границам отрядов (заголовок «Отряд N:»).
    full = "\n".join(body)
    footer = date_label
    if maps:
        max_st = 4 if four else 3
        map_bits = [
            f"{STAGE_ROMAN[i]} {maps[i]}"
            for i in range(min(max_st, len(maps)))
        ]
        footer = f"{date_label} · карты: " + " · ".join(map_bits)

    if len(full) <= 4090:
        embed.description = full
        embed.set_footer(text=footer)
        return embed

    # fallback: куски ≤1024, режем перед «**Отряд» / «**Чемпионы»
    chunks: list[str] = []
    cur: list[str] = []
    cur_n = 0
    for line in body:
        is_break = line.startswith("**") and cur and cur_n > 400
        add = len(line) + (1 if cur else 0)
        if cur and (cur_n + add > 1000 or is_break):
            chunks.append("\n".join(cur))
            cur = []
            cur_n = 0
        cur.append(line)
        cur_n += len(line) + (1 if cur_n else 0)
    if cur:
        chunks.append("\n".join(cur))

    for i, chunk in enumerate(chunks[:20]):
        if not chunk.strip():
            continue
        # имя поля минимальное, чтобы Discord принял; не «·2»
        embed.add_field(name="\u200e", value=chunk, inline=False)  # LTR mark
    embed.set_footer(text=footer + (" · scans/" if len(chunks) > 20 else ""))
    return embed


# ---------------------------------------------------------------------------
# Канал / сообщения / явка
# ---------------------------------------------------------------------------
async def get_log_channel(data: dict[str, Any] | None = None) -> discord.TextChannel | None:
    log_id = get_log_channel_id(data)
    if not log_id:
        return None
    ch = bot.get_channel(log_id)
    if ch is None:
        try:
            ch = await bot.fetch_channel(log_id)
        except Exception:
            return None
    if isinstance(ch, discord.TextChannel):
        return ch
    return None


async def ensure_session_reset(
    data: dict[str, Any],
    *,
    force: bool = False,
) -> dict[str, Any]:
    """
    Открыть сессию КВ (день КВ, после времени явки).
    Гранаты прошлых сессий живут до этого момента.
    force=True — всегда новый старт (тик старта явки / reset_session).
    """
    n = now_msk()
    if not is_cw_day(n):
        return data
    if n.time() < kv_start(n):
        return data

    today = today_msk_str()
    finished = is_kv_finished(data)
    has_scans = bool(data.get("grenade_history"))

    # Уже открыли сегодняшнее окно и КВ ещё не закончен — не трогаем
    if (
        not force
        and data.get("session_date") == today
        and data.get("kv_session_active")
        and not finished
    ):
        return data

    # Сегодня уже сыграли до финала — новую сессию только force
    if (
        not force
        and data.get("session_date") == today
        and finished
    ):
        return data

    # Рестарт mid-KV: есть сканы сегодня, сессия «сброшена» — не стирать, только поднять флаг
    if (
        not force
        and data.get("session_date") == today
        and has_scans
        and not finished
    ):
        if not data.get("kv_session_active"):
            data["kv_session_active"] = True
            save_db(data)
            log("сессия КВ · восстановлена после рестарта (сканы на месте)", "kv")
        return data

    backup_db()
    data["session_date"] = today
    data["grenade_date"] = today
    data["grenade_history"] = {}
    data.pop("combat_history", None)
    data["last_grenade_step"] = None
    data["kv_finished"] = False
    data["kv_session_active"] = True
    data["skipped_grenade_steps"] = []
    data["absent_dm_sent"] = []  # снова можно пингануть неявившимся
    # карты предыдущего дня не тащим; на новый день — чисто
    prev_maps = data.get("kv_maps") or {}
    if not isinstance(prev_maps, dict) or prev_maps.get("date") != today:
        data["kv_maps"] = {"date": None, "maps": []}
    for p in data.get("players", {}).values():
        p["came"] = False
        p["in_voice"] = False
    save_db(data)
    log(
        f"сессия КВ · {WEEKDAYS_RU[now_msk().weekday()]} {now_msk().strftime('%H:%M')} МСК",
        "kv",
    )
    return data


def reset_attendance_only(data: dict[str, Any]) -> dict[str, Any]:
    """При рестарте бота: всем неявка (грены не трогаем)."""
    for p in data.get("players", {}).values():
        p["came"] = False
        p["in_voice"] = False
    return data


def check_channel_post_perms(channel: discord.TextChannel, me: discord.Member) -> str | None:
    """None = ок, иначе текст проблемы."""
    perms = channel.permissions_for(me)
    missing: list[str] = []
    if not perms.view_channel:
        missing.append("Просматривать канал (View Channel)")
    if not perms.send_messages:
        missing.append("Отправлять сообщения (Send Messages)")
    if not perms.embed_links:
        missing.append("Встраивать ссылки (Embed Links)")
    if not perms.read_message_history:
        missing.append("Читать историю сообщений (Read Message History)")
    if missing:
        return "\n".join(f"· {m}" for m in missing)
    return None


# Анти-429: не патчить embeds чаще, чем раз в N секунд (войс-флуд)
_EMBED_MIN_INTERVAL = 12.0
_last_embed_edit_mono: float = 0.0
_embed_flush_task: asyncio.Task | None = None
_embed_pending_parts: set[str] = set()


async def upsert_status_messages(
    data: dict[str, Any] | None = None,
    *,
    force: bool = False,
    parts: tuple[str, ...] | None = None,
) -> tuple[dict[str, Any], str | None]:
    """
    Обновляет/создаёт сообщения ЯВКА и/или ГРАНАТЫ.
    force=True — сразу (скан грен, /setup, /refresh).
    иначе — троттлинг, чтобы не ловить 429 при входе в войсы.
    parts: ("online",), ("grenades",) или оба.
    """
    global _last_embed_edit_mono, _embed_flush_task, _embed_pending_parts

    if data is None:
        data = load_db()

    want = set(parts) if parts else {"online", "grenades"}

    if not force:
        _embed_pending_parts |= want
        now = time.monotonic()
        wait = _EMBED_MIN_INTERVAL - (now - _last_embed_edit_mono)
        if wait > 0:

            async def _delayed_flush() -> None:
                await asyncio.sleep(wait + 0.05)
                async with db_lock:
                    d = load_db()
                    to_send = set(_embed_pending_parts) or {"online"}
                    _embed_pending_parts.clear()
                    await upsert_status_messages(d, force=True, parts=tuple(to_send))

            if _embed_flush_task is None or _embed_flush_task.done():
                _embed_flush_task = asyncio.create_task(_delayed_flush())
            return data, None
        # можно слать сейчас
        want |= _embed_pending_parts
        _embed_pending_parts.clear()

    channel = await get_log_channel(data)
    if channel is None:
        return data, "Канал логов не задан"

    me = channel.guild.me if channel.guild else None
    if me is None and channel.guild:
        me = channel.guild.get_member(bot.user.id) if bot.user else None
    if me is not None:
        bad = check_channel_post_perms(channel, me)
        if bad:
            err = (
                f"Нет прав писать в {channel.mention}.\n"
                f"Выдай боту (или его роли) в этом канале:\n{bad}"
            )
            return data, err

    mids = data.setdefault("message_ids", {"online": None, "grenades": None})

    async def edit_or_send(key: str, embed: discord.Embed) -> int | None:
        mid = mids.get(key)
        if mid:
            try:
                msg = await channel.fetch_message(int(mid))
                await msg.edit(embed=embed)
                return int(mid)
            except (discord.NotFound, discord.Forbidden, discord.HTTPException, TypeError, ValueError):
                pass
        msg = await channel.send(embed=embed)
        return msg.id

    try:
        if "online" in want:
            mids["online"] = await edit_or_send("online", format_online_embed(data))
            await asyncio.sleep(0.35)
        if "grenades" in want:
            mids["grenades"] = await edit_or_send("grenades", format_grenades_embed(data))
    except discord.Forbidden:
        err = (
            f"Нет доступа к каналу {channel.mention} (Missing Access).\n"
            f"Проверь права бота в канале."
        )
        return data, err
    except discord.HTTPException as e:
        if e.status != 429:
            log(f"discord HTTP {e.status}", "warn")
        if not force:
            return data, None
        return data, f"Не удалось отправить сообщение: {e}"

    _last_embed_edit_mono = time.monotonic()
    data["message_ids"] = mids
    save_db(data)
    return data, None


def collect_voice_member_ids(
    guild: discord.Guild,
    voice_ids: list[int] | None = None,
) -> set[int]:
    present: set[int] = set()
    ids = voice_ids if voice_ids is not None else get_voice_channel_ids()
    for vid in ids:
        ch = guild.get_channel(vid)
        if ch is None or not isinstance(ch, discord.VoiceChannel):
            continue
        for m in ch.members:
            if not m.bot:
                present.add(m.id)
    return present


async def refresh_voice_presence() -> None:
    changed = False
    async with db_lock:
        data = load_db()
        guild_id = get_guild_id(data)
        voice_cfg = get_voice_channel_ids(data)
        if not guild_id or not voice_cfg:
            return
        guild = bot.get_guild(guild_id)
        if guild is None:
            return

        data = await ensure_session_reset(data)
        live = is_kv_live(data)
        mark_came = can_mark_came(data)
        voice_ids = collect_voice_member_ids(guild, voice_cfg)

        for did_str, p in data.get("players", {}).items():
            inv = int(did_str) in voice_ids
            # вне окна КВ — 💤, сбрасываем «в войсе»
            if not live:
                if p.get("in_voice"):
                    p["in_voice"] = False
                    changed = True
                continue

            if p.get("in_voice") != inv:
                p["in_voice"] = inv
                changed = True
            # ✅ только до конца сбора явки (20:05)
            if mark_came and inv and not p.get("came"):
                p["came"] = True
                changed = True
                log(f"явка · {p.get('game_nick')}", "ok")

        if changed:
            save_db(data)
    if changed:
        await upsert_status_messages(None, force=False, parts=("online",))


def _step_already_done(last: str | None, step_name: str) -> bool:
    """True, если last_grenade_step уже на этом этапе или позже."""
    if not last or last not in STEP_ORDER or step_name not in STEP_ORDER:
        return False
    return STEP_ORDER.index(last) >= STEP_ORDER.index(step_name)


def _scan_nicks_from_players(players: dict[str, Any]) -> list[str]:
    """Ники для скана: все из отрядов 1–6 + Чемпионы (не «Без отряда»)."""
    nicks: list[str] = []
    seen: set[str] = set()
    for p in players.values():
        if not is_scan_roster_player(p):
            continue
        gn = (p.get("game_nick") or "").strip()
        if not gn:
            continue
        key = gn.lower()
        if key in seen:
            continue
        seen.add(key)
        nicks.append(gn)
    return nicks


def _forward_fill_step(data: dict[str, Any], step_name: str) -> int:
    """
    Пропущенный mid-этап: скопировать последний снапшот gre в step_name.
    Дельта этапа = 0, следующие этапы считаются от корректной базы.
    """
    if step_name not in STEP_ORDER:
        return 0
    idx = STEP_ORDER.index(step_name)
    history = data.setdefault("grenade_history", {})
    n = 0
    for nick, s in list(history.items()):
        if not isinstance(s, dict):
            continue
        if step_name in s:
            continue
        prev_gre = None
        for j in range(idx - 1, -1, -1):
            k = STEP_ORDER[j]
            if k in s and s[k] is not None:
                prev_gre = s[k]
                break
        if prev_gre is None:
            continue
        s[step_name] = prev_gre
        n += 1
    skipped = data.setdefault("skipped_grenade_steps", [])
    if step_name not in skipped:
        skipped.append(step_name)
    return n


async def fill_skipped_grenade_step(step_name: str) -> None:
    """Пометить mid-этап пропущенным (forward-fill), сдвинуть last_step."""
    async with db_lock:
        data = load_db()
        if _step_already_done(data.get("last_grenade_step"), step_name):
            return
        filled = _forward_fill_step(data, step_name)
        data["last_grenade_step"] = step_name
        if not data.get("grenade_date"):
            data["grenade_date"] = data.get("session_date") or today_msk_str()
        save_db(data)
    log(
        f"этап {STEP_TITLES.get(step_name, step_name)} пропущен · fill {filled}",
        "warn",
    )


async def run_grenade_step(step_name: str, prev_step: str | None) -> None:
    """
    Скан gre-thr. API вне db_lock.
    Если API totally fail — этап НЕ закрываем (можно /scan_now).
    """
    step_title = STEP_TITLES.get(step_name, step_name)
    nicks: list[str] = []
    empty_done = False
    data_snap: dict[str, Any] | None = None

    async with db_lock:
        data = load_db()
        if not is_cw_day():
            return

        data = await ensure_session_reset(data)
        final_id = kv_final_step_id(
            kv_dt_from_session(data) if data.get("session_date") else now_msk()
        )
        if is_kv_finished(data) and step_name != final_id:
            return
        if _step_already_done(data.get("last_grenade_step"), step_name):
            return

        nicks = _scan_nicks_from_players(data.get("players") or {})

        if not nicks:
            log(f"грены · {step_title} · нет ников в отрядах 1–6/Чемпионы", "warn")
            data["last_grenade_step"] = step_name
            if step_name == final_id:
                data["kv_finished"] = True
            if not data.get("grenade_date"):
                data["grenade_date"] = data.get("session_date") or today_msk_str()
            save_db(data)
            empty_done = True
            data_snap = data

    if empty_done:
        if data_snap is not None:
            await upsert_status_messages(data_snap, force=True)
        return

    if not nicks:
        return

    # ---- API без lock ----
    results = await scan_grenades(nicks)
    ok_n = sum(1 for gre in results.values() if gre is not None)
    fail_n = len(results) - ok_n

    if ok_n == 0:
        log(
            f"скан · {step_title} · API fail у всех ({fail_n}) — этап НЕ закрыт, повтори /scan_now",
            "err",
        )
        return

    data_for_log: dict[str, Any] | None = None
    async with db_lock:
        data = load_db()
        if _step_already_done(data.get("last_grenade_step"), step_name):
            return
        final_id = kv_final_step_id(
            kv_dt_from_session(data) if data.get("session_date") else now_msk()
        )

        history = data.setdefault("grenade_history", {})
        data.pop("combat_history", None)

        for nick, gre in results.items():
            if gre is None:
                continue
            if nick not in history:
                history[nick] = {}
            history[nick][step_name] = int(gre)

        data["last_grenade_step"] = step_name
        if not data.get("grenade_date"):
            data["grenade_date"] = data.get("session_date") or today_msk_str()
        if step_name == final_id:
            data["kv_finished"] = True
        save_db(data)
        data_for_log = data

    if data_for_log is None:
        return

    await upsert_status_messages(data_for_log, force=True)

    try:
        path = save_grenade_scan(data_for_log, step_name)
        extra = f" · fail {fail_n}" if fail_n else ""
        log(f"скан · {step_title} · ok {ok_n}/{len(nicks)}{extra} · {path.name}", "kv")
    except Exception as e:
        log(f"скан файл: {e}", "err")


def _grenade_raw(s: dict[str, Any], key: str) -> int | None:
    if key not in s or s.get(key) is None:
        return None
    try:
        return int(s.get(key) or 0)
    except (TypeError, ValueError):
        return None


def _stage_diff(a: int | None, b: int | None) -> int | None:
    if a is None or b is None:
        return None
    d = int(a) - int(b)
    return d if d >= 0 else None


def _grenade_stage_values(
    s: dict[str, Any],
) -> tuple:
    """v0..v4 raw + e1..e4 + total."""
    v0 = _grenade_raw(s, "20:00")
    v1 = _grenade_raw(s, "20:25")
    v2 = _grenade_raw(s, "20:50")
    v3 = _grenade_raw(s, "21:20")
    v4 = _grenade_raw(s, "21:40")

    def diff(a: int | None, b: int | None) -> int | None:
        return _stage_diff(a, b)

    def prev_snap(*vals: int | None) -> int | None:
        for v in vals:
            if v is not None:
                return v
        return None

    e1 = diff(v1, v0)
    e2 = diff(v2, prev_snap(v1, v0))
    e3 = diff(v3, prev_snap(v2, v1, v0))
    e4 = diff(v4, prev_snap(v3, v2, v1, v0))
    parts = [x for x in (e1, e2, e3, e4) if x is not None]
    if parts:
        total: int | None = sum(parts)
    elif v0 is not None and all(x is None for x in (v1, v2, v3, v4)):
        total = 0
    else:
        last = prev_snap(v4, v3, v2, v1)
        total = diff(last, v0) if last is not None else None
    return v0, v1, v2, v3, v4, e1, e2, e3, e4, total


def save_grenade_scan(data: dict[str, Any], step_name: str) -> Path:
    """
    Чистый txt-лог:
      scans/etapy/ДАТА_ВРЕМЯ_база|этап1|….txt
      scans/itogi/ДАТА_итог.txt  — после финала
    """
    SCANS_ETAPY_DIR.mkdir(parents=True, exist_ok=True)
    SCANS_ITOGI_DIR.mkdir(parents=True, exist_ok=True)
    history = data.get("grenade_history", {}) or {}
    now = now_msk()
    step_title = STEP_TITLES.get(step_name, step_name)
    slug = GRENADE_STEP_SLUG.get(step_name, step_name.replace(":", ""))
    day = data.get("grenade_date") or data.get("session_date") or now.strftime("%Y-%m-%d")
    time_part = now.strftime("%H%M")
    out = SCANS_ETAPY_DIR / f"{day}_{time_part}_{slug}.txt"

    players = data.get("players", {})
    nick_meta: dict[str, tuple] = {}
    for p in players.values():
        gn = p.get("game_nick")
        if gn:
            nick_meta[gn] = (p.get("squad"), p.get("slot"), gn.lower())
            nick_meta[gn.lower()] = nick_meta[gn]

    def sort_key(nick: str) -> tuple:
        meta = nick_meta.get(nick) or nick_meta.get(nick.lower())
        if meta:
            sq, sl, low = meta
            has = 0 if sq is not None else 1
            try:
                sq_n = int(sq) if sq is not None else 999
            except (TypeError, ValueError):
                sq_n = 999
            try:
                sl_n = int(sl) if sl is not None else 999
            except (TypeError, ValueError):
                sl_n = 999
            return (has, sq_n, sl_n, low)
        return (1, 999, 999, nick.lower())

    nicks_sorted = sorted(history.keys(), key=sort_key)

    try:
        y, m, d0 = map(int, str(day).split("-")[:3])
        day_dt = datetime(y, m, d0, tzinfo=MSK)
    except Exception:
        day_dt = now
    four = _weekday(day_dt) == 6
    maps = get_kv_maps(data, for_date=str(day))
    cur_map = map_for_step(data, step_name)

    lines: list[str] = [
        f"{day}  {WEEKDAYS_RU[day_dt.weekday()]}  {now.strftime('%H:%M')} МСК",
        f"{step_title}" + (f"  ·  {cur_map}" if cur_map else ""),
        f"игроков: {len(history)}",
    ]
    if maps:
        max_st = 4 if four else 3
        map_bits = [
            f"{STAGE_ROMAN[i]} {maps[i]}"
            for i in range(min(max_st, len(maps)))
        ]
        lines.append("карты: " + " · ".join(map_bits))
    lines.append("")

    cw = 5
    nick_w = 16
    col_gap = "  "
    tot_gap = "     "

    def _cols(*cells: str, with_total: str | None = None) -> str:
        mid = col_gap.join(cells)
        if with_total is None:
            return mid
        return mid + tot_gap + with_total

    if four:
        cols = ("I", "II", "III", "IV", "ИТОГ")
    else:
        cols = ("I", "II", "III", "ИТОГ")
    hdr_cells = [f"{c:>{cw}}" for c in cols[:-1]]
    hdr = f"{'ник':<{nick_w}} {_cols(*hdr_cells, with_total=f'{cols[-1]:>{cw}}')}"
    lines.append(hdr)
    lines.append("-" * len(hdr))

    last_sq: Any = object()
    for nick in nicks_sorted:
        meta = nick_meta.get(nick) or nick_meta.get(nick.lower())
        sq = meta[0] if meta else None
        if sq != last_sq:
            if last_sq is not object():
                lines.append("")
            last_sq = sq
            lines.append(squad_label(data, sq))

        e1, e2, e3, e4, total = _grenade_row_stats(history, nick, four_stages=four)
        nick_show = nick if len(nick) <= nick_w else nick[: nick_w - 1] + "…"
        if four:
            line = (
                f"{nick_show:<{nick_w}} "
                f"{_cols(cell(e1, cw), cell(e2, cw), cell(e3, cw), cell(e4, cw), with_total=cell(total, cw))}"
            )
        else:
            line = (
                f"{nick_show:<{nick_w}} "
                f"{_cols(cell(e1, cw), cell(e2, cw), cell(e3, cw), with_total=cell(total, cw))}"
            )
        lines.append(line)

    lines.append("")
    text = "\n".join(lines) + "\n"
    out.write_text(text, encoding="utf-8")

    if step_name == kv_final_step_id(day_dt):
        summary = SCANS_ITOGI_DIR / f"{day}_итог.txt"
        itog_lines = lines.copy()
        if itog_lines:
            itog_lines[0] = itog_lines[0] + "  ·  итог"
        summary.write_text("\n".join(itog_lines) + "\n", encoding="utf-8")
        log(f"итог · itogi/{summary.name}", "ok")

    return out


# ---------------------------------------------------------------------------
# Slash-команды · настройка каналов
# ---------------------------------------------------------------------------
def format_settings_embed(data: dict[str, Any]) -> discord.Embed:
    cfg = ensure_config(data)
    log_id = get_log_channel_id(data)
    voices = get_voice_channel_ids(data)
    guild_id = get_guild_id(data)

    log_txt = f"<#{log_id}>" if log_id else "❌ не задан"
    if voices:
        voice_txt = "\n".join(f"· <#{vid}>" for vid in voices)
    else:
        voice_txt = "❌ не заданы"

    ok_log = "✅" if log_id else "❌"
    ok_voice = "✅" if voices else "❌"
    ready = bool(log_id and voices)
    access_ids = get_access_role_ids(data)
    access_txt = "\n".join(f"· <@&{rid}>" for rid in access_ids) if access_ids else "❌ не заданы"

    embed = discord.Embed(
        title="⚙️  НАСТРОЙКИ БОТА",
        description=(
            f"**Статус:** {'🟢 готов к работе' if ready else '🟡 нужно донастроить'}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"{ok_log} **Канал логов:** {log_txt}\n"
            f"{ok_voice} **Войс-каналы КВ:**\n{voice_txt}\n"
            f"🔑 **Роли /add /remove:**\n{access_txt}\n"
            f"\n🏠 Сервер ID: `{guild_id or '—'}`"
        ),
        color=COLOR_OK if ready else COLOR_WAIT,
        timestamp=now_msk(),
    )
    embed.add_field(
        name="Как настроить",
        value=(
            "**/setup** — канал логов и войсы\n"
            "**/access_add** — роль для `/add` `/remove` (полный доступ)\n"
            "**/access_remove** · **/access_list**\n"
            "**/settings** — эти настройки"
        ),
        inline=False,
    )
    embed.set_footer(text="КВ STALCRAFT · настройка")
    return embed


@bot.tree.command(name="setup", description="Настроить канал логов и войс-каналы КВ")
@app_commands.describe(
    log_channel="Текстовый канал, куда бот пишет явку и гранаты",
    voice1="Войс-канал №1 (обязательно хотя бы один)",
    voice2="Войс-канал №2 (необязательно)",
    voice3="Войс-канал №3 (необязательно)",
    voice4="Войс-канал №4 (необязательно)",
    voice5="Войс-канал №5 (необязательно)",
)
async def cmd_setup(
    interaction: discord.Interaction,
    log_channel: discord.TextChannel,
    voice1: discord.VoiceChannel,
    voice2: discord.VoiceChannel | None = None,
    voice3: discord.VoiceChannel | None = None,
    voice4: discord.VoiceChannel | None = None,
    voice5: discord.VoiceChannel | None = None,
):
    if interaction.guild is None:
        await interaction.response.send_message(
            embed=make_reply_embed("❌  Ошибка", "Команда только на сервере.", COLOR_ERR),
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)

    voices: list[discord.VoiceChannel] = []
    for v in (voice1, voice2, voice3, voice4, voice5):
        if v is not None and v.id not in {x.id for x in voices}:
            voices.append(v)

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        old_log = cfg.get("log_channel_id")
        cfg["guild_id"] = interaction.guild.id
        cfg["log_channel_id"] = log_channel.id
        cfg["voice_channel_ids"] = [v.id for v in voices]
        if old_log and int(old_log) != log_channel.id:
            data["message_ids"] = {"online": None, "grenades": None}
        save_db(data)
        _data, post_err = await upsert_status_messages(data)

    voice_list = "\n".join(f"· {v.mention}" for v in voices)
    if post_err:
        await interaction.followup.send(
            embed=make_reply_embed(
                "⚠️  Настройки сохранены, но писать в канал нельзя",
                (
                    f"**Канал логов:** {log_channel.mention}\n"
                    f"**Войсы КВ:**\n{voice_list}\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"{post_err}\n\n"
                    f"После выдачи прав нажми `/refresh`."
                ),
                color=COLOR_LIVE,
            ),
            ephemeral=True,
        )
        return

    await interaction.followup.send(
        embed=make_reply_embed(
            "✅  Настройка сохранена",
            (
                f"**Канал логов:** {log_channel.mention}\n"
                f"**Войсы КВ:**\n{voice_list}\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"Сообщения **ЯВКА** и **ГРАНАТЫ** отправлены в канал."
            ),
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="set_log", description="Указать текстовый канал для логов бота")
@app_commands.describe(channel="Куда писать явку и гранаты")
async def cmd_set_log(interaction: discord.Interaction, channel: discord.TextChannel):
    if interaction.guild is None:
        await interaction.response.send_message(
            embed=make_reply_embed("❌  Ошибка", "Команда только на сервере.", COLOR_ERR),
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        old_log = cfg.get("log_channel_id")
        cfg["guild_id"] = interaction.guild.id
        cfg["log_channel_id"] = channel.id
        if old_log and int(old_log) != channel.id:
            data["message_ids"] = {"online": None, "grenades": None}
        save_db(data)
        _data, post_err = await upsert_status_messages(data)

    if post_err:
        await interaction.followup.send(
            embed=make_reply_embed(
                "⚠️  Канал сохранён, но писать нельзя",
                f"{channel.mention}\n\n{post_err}\n\nПосле прав — `/refresh`.",
                color=COLOR_LIVE,
            ),
            ephemeral=True,
        )
        return

    await interaction.followup.send(
        embed=make_reply_embed(
            "✅  Канал логов задан",
            f"Бот пишет сюда: {channel.mention}\n"
            f"Не забудь войсы: `/voice_add` или `/setup`.",
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="voice_add", description="Добавить войс-канал для отслеживания явки")
@app_commands.describe(channel="Войс, в котором отмечаем «пришёл»")
async def cmd_voice_add(interaction: discord.Interaction, channel: discord.VoiceChannel):
    if interaction.guild is None:
        await interaction.response.send_message(
            embed=make_reply_embed("❌  Ошибка", "Команда только на сервере.", COLOR_ERR),
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        cfg["guild_id"] = interaction.guild.id
        ids = [int(x) for x in (cfg.get("voice_channel_ids") or [])]
        if channel.id in ids:
            await interaction.followup.send(
                embed=make_reply_embed(
                    "ℹ️  Уже в списке",
                    f"{channel.mention} уже отслеживается.",
                    color=COLOR_WAIT,
                ),
                ephemeral=True,
            )
            return
        ids.append(channel.id)
        cfg["voice_channel_ids"] = ids
        save_db(data)
        await upsert_status_messages(data)

    await interaction.followup.send(
        embed=make_reply_embed(
            "✅  Войс добавлен",
            f"Теперь слежу за: {channel.mention}\n"
            f"Всего войсов: **{len(ids)}**",
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="voice_remove", description="Убрать войс-канал из отслеживания")
@app_commands.describe(channel="Какой войс больше не считать")
async def cmd_voice_remove(interaction: discord.Interaction, channel: discord.VoiceChannel):
    await interaction.response.defer(ephemeral=True)

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        ids = [int(x) for x in (cfg.get("voice_channel_ids") or [])]
        if channel.id not in ids:
            if not ids and ENV_VOICE_CHANNEL_IDS:
                ids = list(ENV_VOICE_CHANNEL_IDS)
            if channel.id not in ids:
                await interaction.followup.send(
                    embed=make_reply_embed(
                        "ℹ️  Не найден",
                        f"{channel.mention} нет в списке войсов.",
                        color=COLOR_WAIT,
                    ),
                    ephemeral=True,
                )
                return
        ids = [x for x in ids if x != channel.id]
        cfg["voice_channel_ids"] = ids
        if interaction.guild:
            cfg["guild_id"] = interaction.guild.id
        save_db(data)
        await upsert_status_messages(data)

    await interaction.followup.send(
        embed=make_reply_embed(
            "🗑️  Войс убран",
            f"{channel.mention} больше не отслеживается.\n"
            f"Осталось войсов: **{len(ids)}**",
            color=COLOR_ERR,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="voice_clear", description="Очистить весь список войс-каналов")
async def cmd_voice_clear(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        cfg["voice_channel_ids"] = []
        if interaction.guild:
            cfg["guild_id"] = interaction.guild.id
        save_db(data)
        await upsert_status_messages(data)

    await interaction.followup.send(
        embed=make_reply_embed(
            "🧹  Список войсов очищен",
            "Добавь каналы снова через `/voice_add` или `/setup`.",
            color=COLOR_INFO,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="settings", description="Показать текущие настройки каналов")
async def cmd_settings(interaction: discord.Interaction):
    data = load_db()
    await interaction.response.send_message(
        embed=format_settings_embed(data),
        ephemeral=True,
    )


def load_phrases_file(path: Path) -> list[str]:
    if not path.exists():
        return []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except Exception:
        return []
    out: list[str] = []
    for line in lines:
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        out.append(s)
    return out


async def resolve_user_for_dm(
    guild: discord.Guild | None,
    uid: int,
) -> discord.Member | discord.User | None:
    if guild is not None:
        m = guild.get_member(uid)
        if m is not None:
            return m
        try:
            return await guild.fetch_member(uid)
        except Exception:
            pass
    try:
        return await bot.fetch_user(uid)
    except Exception:
        return None


async def dm_never_came_players(
    guild: discord.Guild | None,
    *,
    force_resend: bool = False,
    fixed_text: str | None = None,
) -> dict[str, Any]:
    """
    ЛС тем, у кого came=False с начала сессии.
    Текст (из txt или fixed_text), затем пинг.
    Один раз за сессию (absent_dm_sent), если force_resend=False.
    """
    async with db_lock:
        data = load_db()
        players = data.get("players") or {}
        sent_list = [str(x) for x in (data.get("absent_dm_sent") or [])]
        sent_set = set(sent_list)

        targets: list[tuple[str, dict[str, Any]]] = []
        skip_subs = 0
        for did_str, p in players.items():
            if p.get("came"):
                continue
            # Чемпионы (замены) — не пишем в ЛС перед КВ
            if is_subs_player(p):
                skip_subs += 1
                continue
            if not force_resend and did_str in sent_set:
                continue
            targets.append((did_str, p))

    phrases = load_phrases_file(ABSENT_DM_PHRASES_FILE)
    default_phrase = "ты ещё не заходил в войс на КВ — заходи, пока окно явки открыто"

    ok = 0
    fail = 0
    skip_already = 0
    fail_lines: list[str] = []
    newly_sent: list[str] = []

    for did_str, p in targets:
        try:
            uid = int(did_str)
        except ValueError:
            fail += 1
            fail_lines.append(f"· плохой id `{did_str}`")
            continue

        member = await resolve_user_for_dm(guild, uid)
        if member is None:
            fail += 1
            fail_lines.append(f"· <@{did_str}> (`{p.get('game_nick')}`) — не найден")
            continue

        phrase = (fixed_text.strip() if fixed_text and fixed_text.strip() else None) or (
            random.choice(phrases) if phrases else default_phrase
        )
        if len(phrase) > 1800:
            phrase = phrase[:1800] + "…"
        # текст, потом пинг
        msg = f"{phrase}\n{member.mention}"

        try:
            await member.send(msg)
            ok += 1
            newly_sent.append(did_str)
        except discord.Forbidden:
            fail += 1
            fail_lines.append(
                f"· {member.mention} (`{p.get('game_nick')}`) — закрыты ЛС"
            )
        except discord.HTTPException as e:
            fail += 1
            fail_lines.append(f"· {member.mention} — HTTP: {e}")

        await asyncio.sleep(0.7)

    if newly_sent:
        async with db_lock:
            data = load_db()
            cur = [str(x) for x in (data.get("absent_dm_sent") or [])]
            for did in newly_sent:
                if did not in cur:
                    cur.append(did)
            data["absent_dm_sent"] = cur
            save_db(data)

    # сколько пропущено т.к. уже слали
    async with db_lock:
        data = load_db()
        players = data.get("players") or {}
        sent_set = set(str(x) for x in (data.get("absent_dm_sent") or []))
        for did_str, p in players.items():
            if not p.get("came") and did_str in sent_set and did_str not in newly_sent:
                if not force_resend:
                    skip_already += 1

    return {
        "ok": ok,
        "fail": fail,
        "skip_already": skip_already,
        "skip_subs": skip_subs,
        "targets": len(targets),
        "fail_lines": fail_lines,
        "never_came_total": sum(
            1
            for p in (load_db().get("players") or {}).values()
            if not p.get("came") and not is_subs_player(p)
        ),
    }


@bot.tree.command(
    name="dm_absent",
    description="ЛС тем, кто ни разу не заходил в войс с начала явки",
)
@app_commands.describe(
    text="Свой текст (иначе случайный из absent_dm_phrases.txt)",
    force="Слать даже тем, кому уже писали в этой сессии",
)
@app_commands.guild_only()
async def cmd_dm_absent(
    interaction: discord.Interaction,
    text: str | None = None,
    force: bool = False,
):
    actor = await resolve_member(interaction)
    if not is_bot_admin(actor):
        await interaction.response.send_message(
            embed=make_reply_embed(
                "❌  Нет прав",
                "Только админ / Manage Server.",
                color=COLOR_ERR,
            ),
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)
    result = await dm_never_came_players(
        interaction.guild,
        force_resend=force,
        fixed_text=text,
    )

    report = (
        f"Не было в войсе (без замен): **{result['never_came_total']}**\n"
        f"К отправке сейчас: **{result['targets']}**\n"
        f"✅ Доставлено: **{result['ok']}**\n"
        f"⏭ Уже писали (сессия): **{result['skip_already']}**\n"
        f"🛡 Пропуск Чемпионы/замены: **{result.get('skip_subs', 0)}**\n"
        f"❌ Не удалось: **{result['fail']}**\n\n"
        f"_Формат: текст → пинг. Один раз за сессию (если не force)._\n"
        f"_Чемпионы (замены) не получают ЛС._"
    )
    if result["fail_lines"]:
        chunk = "\n".join(result["fail_lines"][:15])
        if len(result["fail_lines"]) > 15:
            chunk += f"\n… ещё {len(result['fail_lines']) - 15}"
        report += f"\n\n**Ошибки:**\n{chunk}"

    await interaction.followup.send(
        embed=make_reply_embed(
            "📬  ЛС неявившимся",
            report,
            color=COLOR_INFO if result["ok"] or result["targets"] == 0 else COLOR_ERR,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="say", description="Написать сообщение от лица бота в канал")
@app_commands.describe(
    channel="Куда отправить (текстовый канал)",
    text="Текст сообщения",
)
@app_commands.guild_only()
async def cmd_say(
    interaction: discord.Interaction,
    channel: discord.TextChannel,
    text: str,
):
    actor = await resolve_member(interaction)
    if not is_bot_admin(actor):
        await interaction.response.send_message(
            embed=make_reply_embed(
                "❌  Нет прав",
                "Писать от бота могут только админы сервера (или Manage Server).",
                color=COLOR_ERR,
            ),
            ephemeral=True,
        )
        return

    text = text.strip()
    if not text:
        await interaction.response.send_message(
            embed=make_reply_embed("❌  Ошибка", "Пустой текст.", color=COLOR_ERR),
            ephemeral=True,
        )
        return
    if len(text) > 2000:
        await interaction.response.send_message(
            embed=make_reply_embed(
                "❌  Слишком длинно",
                "Лимит Discord — **2000** символов.",
                color=COLOR_ERR,
            ),
            ephemeral=True,
        )
        return

    await interaction.response.defer(ephemeral=True)

    me = channel.guild.me if channel.guild else None
    if me is not None:
        perms = channel.permissions_for(me)
        if not perms.view_channel or not perms.send_messages:
            await interaction.followup.send(
                embed=make_reply_embed(
                    "❌  Нет прав у бота",
                    f"В {channel.mention} боту нужны **Просматривать канал** и **Отправлять сообщения**.",
                    color=COLOR_ERR,
                ),
                ephemeral=True,
            )
            return

    try:
        await channel.send(content=text)
    except discord.Forbidden:
        await interaction.followup.send(
            embed=make_reply_embed(
                "❌  Forbidden",
                f"Бот не может писать в {channel.mention}.",
                color=COLOR_ERR,
            ),
            ephemeral=True,
        )
        return
    except discord.HTTPException as e:
        await interaction.followup.send(
            embed=make_reply_embed("❌  Ошибка Discord", str(e), color=COLOR_ERR),
            ephemeral=True,
        )
        return

    await interaction.followup.send(
        embed=make_reply_embed(
            "✅  Отправлено",
            f"В {channel.mention}:\n>>> {text[:500]}{'…' if len(text) > 500 else ''}",
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


# ---------------------------------------------------------------------------
# Доступ ролей к /add /remove
# ---------------------------------------------------------------------------
@bot.tree.command(name="access_add", description="Выдать роли доступ к /add и /remove")
@app_commands.describe(role="Роль, например @дитя")
async def cmd_access_add(interaction: discord.Interaction, role: discord.Role):
    if not isinstance(interaction.user, discord.Member) or not is_bot_admin(interaction.user):
        await interaction.response.send_message(
            embed=make_reply_embed(
                "❌  Нет прав",
                "Только администратор или Manage Server может выдавать доступ.",
                color=COLOR_ERR,
            ),
            ephemeral=True,
        )
        return

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        ids = [int(x) for x in (cfg.get("access_role_ids") or [])]
        if role.id in ids:
            await interaction.response.send_message(
                embed=make_reply_embed(
                    "ℹ️  Уже есть",
                    f"{role.mention} уже имеет доступ к `/add` и `/remove`.",
                    color=COLOR_WAIT,
                ),
                ephemeral=True,
            )
            return
        ids.append(role.id)
        cfg["access_role_ids"] = ids
        if interaction.guild:
            cfg["guild_id"] = interaction.guild.id
        save_db(data)

    await interaction.response.send_message(
        embed=make_reply_embed(
            "✅  Доступ выдан",
            (
                f"Роль {role.mention} — полный доступ к составу:\n"
                f"· `/add` · `/remove` (как админ)\n"
                f"· или в чат: `!add ник` / `!remove @user` / `!remove ник`\n\n"
                f"**Если /add не видно** — Настройки сервера → Интеграции → "
                f"далбаебик → add/remove → включи {role.mention} или @everyone."
            ),
            color=COLOR_OK,
        ),
    )


@bot.tree.command(name="access_remove", description="Забрать у роли доступ к /add и /remove")
@app_commands.describe(role="Роль, у которой убрать доступ")
async def cmd_access_remove(interaction: discord.Interaction, role: discord.Role):
    if not isinstance(interaction.user, discord.Member) or not is_bot_admin(interaction.user):
        await interaction.response.send_message(
            embed=make_reply_embed(
                "❌  Нет прав",
                "Только администратор или Manage Server.",
                color=COLOR_ERR,
            ),
            ephemeral=True,
        )
        return

    async with db_lock:
        data = load_db()
        cfg = ensure_config(data)
        ids = [int(x) for x in (cfg.get("access_role_ids") or [])]
        if role.id not in ids:
            await interaction.response.send_message(
                embed=make_reply_embed(
                    "ℹ️  Не найдено",
                    f"{role.mention} нет в списке доступа.",
                    color=COLOR_WAIT,
                ),
                ephemeral=True,
            )
            return
        cfg["access_role_ids"] = [x for x in ids if x != role.id]
        save_db(data)

    await interaction.response.send_message(
        embed=make_reply_embed(
            "🗑️  Доступ снят",
            f"Роль {role.mention} больше не может `/add` / `/remove`.",
            color=COLOR_ERR,
        ),
    )


@bot.tree.command(name="access_list", description="Список ролей с доступом к /add и /remove")
async def cmd_access_list(interaction: discord.Interaction):
    data = load_db()
    ids = get_access_role_ids(data)
    if not ids:
        text = "_Ролей нет. Админ: `/access_add @роль`_"
    else:
        text = "\n".join(f"· <@&{rid}>" for rid in ids)
    await interaction.response.send_message(
        embed=make_reply_embed(
            "🔑  Доступ к составу",
            (
                f"{text}\n\n"
                f"Эти роли + админы: полный `/add` и `/remove` (кого угодно)."
            ),
            color=COLOR_INFO,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="add", description="Добавить игрока в базу (Discord + ник STALCRAFT)")
@app_commands.describe(
    game_nick="Ник в STALCRAFT",
    user="Кого добавить (пустым = себя)",
)
@app_commands.guild_only()
async def cmd_add(
    interaction: discord.Interaction,
    game_nick: str,
    user: discord.Member | None = None,
):
    await interaction.response.defer(ephemeral=False)
    actor = await resolve_member(interaction)
    target = user if user is not None else actor
    if actor is None or target is None:
        await reply_interaction(
            interaction,
            embed=make_reply_embed("❌  Ошибка", "Не удалось определить участника.", color=COLOR_ERR),
            ephemeral=True,
        )
        return
    ok, emb = await do_roster_add(actor, target, game_nick)
    await reply_interaction(interaction, embed=emb, ephemeral=False)
    if ok:
        schedule_add_phrase(interaction.channel)


@bot.tree.command(name="remove", description="Удалить из базы: @Discord или ник STALCRAFT")
@app_commands.describe(
    user="Discord-участник (пустым + без ника = себя)",
    game_nick="Или ник STALCRAFT из таблицы (sosew, Teipo…)",
)
@app_commands.guild_only()
async def cmd_remove(
    interaction: discord.Interaction,
    user: discord.Member | None = None,
    game_nick: str | None = None,
):
    await interaction.response.defer(ephemeral=False)
    actor = await resolve_member(interaction)
    if actor is None:
        await reply_interaction(
            interaction,
            embed=make_reply_embed("❌  Ошибка", "Не удалось определить участника.", color=COLOR_ERR),
            ephemeral=True,
        )
        return
    nick = (game_nick or "").strip() or None
    if user is None and nick is None:
        # как раньше: без аргументов = себя
        emb = await do_roster_remove(actor, target=actor)
    else:
        emb = await do_roster_remove(actor, target=user, game_nick=nick)
    await reply_interaction(interaction, embed=emb, ephemeral=False)


# Запасной вариант, если Discord скрыл slash у роли
@bot.command(name="add")
async def prefix_add(ctx: commands.Context, game_nick: str, member: discord.Member | None = None):
    """!add ник_stalcraft   или   !add ник @user"""
    if not isinstance(ctx.author, discord.Member):
        await ctx.reply("Только на сервере.")
        return
    target = member or ctx.author
    ok, emb = await do_roster_add(ctx.author, target, game_nick)
    await ctx.reply(embed=emb)
    if ok:
        schedule_add_phrase(ctx.channel)


@bot.command(name="remove")
async def prefix_remove(ctx: commands.Context, *, who: str | None = None):
    """!remove  |  !remove @user  |  !remove sosew"""
    if not isinstance(ctx.author, discord.Member):
        await ctx.reply("Только на сервере.")
        return
    if who is None or not str(who).strip():
        emb = await do_roster_remove(ctx.author, target=ctx.author)
        await ctx.reply(embed=emb)
        return

    raw = str(who).strip()
    # сначала пробуем как @Discord / id / имя
    try:
        member = await commands.MemberConverter().convert(ctx, raw)
        emb = await do_roster_remove(ctx.author, target=member)
        await ctx.reply(embed=emb)
        return
    except commands.BadArgument:
        pass

    # иначе — ник STALCRAFT
    emb = await do_roster_remove(ctx.author, game_nick=raw)
    await ctx.reply(embed=emb)


async def do_set_maps(
    actor: discord.Member,
    maps_text: str | None,
) -> tuple[bool, discord.Embed]:
    """
    /map · !map
    maps_text=None → показать текущие
    maps_text=clear/сброс → очистить
    иначе → задать 3–4 карты на этапы I–III(IV)
    """
    data_preview = load_db()
    if not can_use_roster(actor, data_preview):
        allowed = get_access_role_ids(data_preview)
        roles_txt = ", ".join(f"<@&{r}>" for r in allowed) if allowed else "_ролей нет — только админ_"
        return False, make_reply_embed(
            "❌  Нет доступа",
            f"Нужна роль: {roles_txt}\nили админ.",
            color=COLOR_ERR,
        )

    four = _weekday() == 6
    max_st = 4 if four else 3

    # показать
    if maps_text is None or not str(maps_text).strip():
        maps = get_kv_maps(data_preview)
        body = format_maps_lines(maps, max_stages=max_st)
        hint = (
            f"\n\nЗадать: `/map хвойник берда низина`"
            f"{' [карта IV]' if four else ''}"
            f"\nСброс: `/map clear`"
        )
        return True, make_reply_embed(
            "🗺️  Карты КВ",
            body + hint,
            color=COLOR_INFO if maps else COLOR_WAIT,
        )

    text = str(maps_text).strip()
    low = text.lower().replace("ё", "е")
    if low in ("clear", "сброс", "очистить", "reset", "none", "-", "0"):
        async with db_lock:
            data = load_db()
            clear_kv_maps(data)
            save_db(data)
            await upsert_status_messages(data, force=True, parts=("grenades",))
        return True, make_reply_embed(
            "🗺️  Карты сброшены",
            "Подписи этапов в **ГРАНАТЫ** убраны.",
            color=COLOR_OK,
        )

    maps, err = parse_maps_text(text, expect=None)
    if err or not maps:
        return False, make_reply_embed("❌  Карты", err or "Не разобрал.", color=COLOR_ERR)

    # в вс можно 3 или 4; в чт–сб — ровно 3 (4-ю молча отрежем с предупреждением)
    note = ""
    if four:
        if len(maps) not in (3, 4):
            return False, make_reply_embed(
                "❌  Карты",
                f"В **воскресенье** нужно **3 или 4** карты (сейчас {len(maps)}).",
                color=COLOR_ERR,
            )
    else:
        if len(maps) < 3:
            return False, make_reply_embed(
                "❌  Карты",
                f"Нужно **3** карты на этапы I–III (сейчас {len(maps)}).",
                color=COLOR_ERR,
            )
        if len(maps) > 3:
            note = f"\n_чт–сб: 3 этапа — 4-ю карту (`{maps[3]}`) не сохранил._"
            maps = maps[:3]

    async with db_lock:
        data = load_db()
        day = data.get("grenade_date") or data.get("session_date") or today_msk_str()
        set_kv_maps(data, maps, day=day)
        save_db(data)
        await upsert_status_messages(data, force=True, parts=("grenades",))

    body = format_maps_lines(maps, max_stages=max_st) + note
    log(f"карты · {' · '.join(f'{STAGE_ROMAN[i]} {m}' for i, m in enumerate(maps))}", "kv")
    return True, make_reply_embed(
        "✅  Карты закреплены",
        body + "\n\nПоявятся в **ГРАНАТЫ** и в txt-сканах.",
        color=COLOR_OK,
    )


@bot.tree.command(name="map", description="Карты КВ на этапы I–III (вс: +IV)")
@app_commands.describe(
    maps="3 карты: хвойник берда низина  ·  clear — сброс  ·  пусто — показать",
)
@app_commands.guild_only()
async def cmd_map(
    interaction: discord.Interaction,
    maps: str | None = None,
):
    await interaction.response.defer(ephemeral=False)
    actor = await resolve_member(interaction)
    if actor is None:
        await reply_interaction(
            interaction,
            embed=make_reply_embed("❌  Ошибка", "Не удалось определить участника.", color=COLOR_ERR),
            ephemeral=True,
        )
        return
    _ok, emb = await do_set_maps(actor, maps)
    await reply_interaction(interaction, embed=emb, ephemeral=False)


@bot.command(name="map")
async def prefix_map(ctx: commands.Context, *, maps: str | None = None):
    """!map хвойник берда низина   ·   !map   ·   !map clear"""
    if not isinstance(ctx.author, discord.Member):
        await ctx.reply("Только на сервере.")
        return
    _ok, emb = await do_set_maps(ctx.author, maps)
    await ctx.reply(embed=emb)


@bot.tree.command(
    name="fix_slash",
    description="Админ: пересинхронизировать slash + инструкция если /add не виден",
)
@app_commands.guild_only()
async def cmd_fix_slash(interaction: discord.Interaction):
    actor = await resolve_member(interaction)
    if not is_bot_admin(actor):
        await interaction.response.send_message(
            embed=make_reply_embed("❌  Нет прав", "Только админ сервера.", color=COLOR_ERR),
            ephemeral=True,
        )
        return
    await interaction.response.defer(ephemeral=True)
    if interaction.guild is None:
        await interaction.followup.send("Только на сервере.", ephemeral=True)
        return
    for name in ("add", "remove"):
        cmd = bot.tree.get_command(name)
        if cmd is not None:
            try:
                cmd.default_permissions = None
            except Exception:
                pass
    bot.tree.copy_global_to(guild=interaction.guild)
    synced = await bot.tree.sync(guild=interaction.guild)
    await interaction.followup.send(
        embed=make_reply_embed(
            f"✅  Синхронизировано ({len(synced)} команд)",
            (
                "Discord **не даёт ботам** включать slash для ролей (ошибка 403).\n\n"
                "**Сделай руками (1 раз):**\n"
                "1. Настройки сервера → **Интеграции** → **далбаебик**\n"
                "2. Команды **add** и **remove**\n"
                "3. Включи для **@everyone** или **@дитя**\n\n"
                "**Сейчас точно работает в чате:**\n"
                "`!add ник_stalcraft`\n"
                "`!add ник_stalcraft @user`\n"
                "`!remove` / `!remove @user` / `!remove sosew`\n\n"
                "В Portal должен быть включён **Message Content Intent**."
            ),
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="help", description="Список команд бота")
async def cmd_help(interaction: discord.Interaction):
    text = (
        "**Состав**\n"
        "· `/add` · `!add` — Discord + ник STALCRAFT\n"
        "· `/remove` · `!remove` — убрать: @Discord **или** ник STALCRAFT\n"
        "· `/list` — весь состав\n"
        "\n"
        "**Отряды (Excel)**\n"
        "· `/sheet_sync` — прочитать таблицу сейчас\n"
        "· `/squad_list` — сетка отрядов\n"
        "· `/sheet_path` — путь к Excel\n"
        "\n"
        "**КВ / явка / гранаты**\n"
        "· **чт–сб:** явка 19:30–20:05 · 3 этапа · 20:05/25/50/21:20\n"
        "· **вс:** явка 18:30–19:00 · **4 этапа** · 19:00/20/40 · 20:00/20:20\n"
        "· `/map` · `!map` — карты на этапы\n"
        "· скан грен: **отряды 1–6 + Чемпионы** (не войс)\n"
        "· данные грен висят до следующего **19:30** (обнуление сессии)\n"
        "· `/refresh` · `/scan_now`* · `/deletegren`* · `/reset_session`*\n"
        "· `/dm_absent` — ЛС (без Чемпионов/замен)\n"
        "· \\* = access-роль или админ\n"
        "\n"
        "**Настройка (админ)**\n"
        "· `/setup` — канал логов + войсы\n"
        "· `/set_log` · `/voice_add` · `/voice_remove` · `/voice_clear`\n"
        "· `/settings` — текущие настройки\n"
        "· `/access_add` · `/access_remove` · `/access_list` — кто может `/add`\n"
        "· `/say` — сообщение от бота\n"
        "\n"
        f"Excel: `{SHEET_PATH.name}` · автосинк раз в {SHEET_SYNC_SECONDS} сек"
    )
    await interaction.response.send_message(
        embed=make_reply_embed("📖  Команды бота", text, color=COLOR_INFO),
        ephemeral=True,
    )


@bot.tree.command(name="list", description="Показать весь состав из базы")
async def cmd_list(interaction: discord.Interaction):
    data = load_db()
    players = data.get("players", {})
    if not players:
        await interaction.response.send_message(
            embed=make_reply_embed(
                "👥  Состав",
                "База пуста. Добавь игроков через `/add`.",
                color=COLOR_WAIT,
            ),
            ephemeral=True,
        )
        return

    lines = []
    for i, (did, p) in enumerate(
        sorted(players.items(), key=lambda x: (x[1].get("discord_name") or "").lower()),
        start=1,
    ):
        came = "✅" if p.get("came") else "❌"
        voice = "🔊" if p.get("in_voice") else "🔇"
        lines.append(
            f"`{i:>2}.` {came}{voice} `{p.get('game_nick')}` | <@{did}>"
        )

    embed = discord.Embed(
        title="👥  СОСТАВ КЛАНА",
        description=(
            f"Всего в базе: **{len(players)}**\n"
            f"━━━━━━━━━━━━━━━━━━━━"
        ),
        color=COLOR_INFO,
        timestamp=now_msk(),
    )
    for i, chunk in enumerate(split_field_chunks(lines)):
        embed.add_field(
            name="Игроки" if i == 0 else f"Игроки · {i + 1}",
            value=chunk,
            inline=False,
        )
    embed.set_footer(text="КВ STALCRAFT · только для тебя")
    await interaction.response.send_message(embed=embed, ephemeral=True)


@bot.tree.command(name="refresh", description="Принудительно обновить сообщения явки и гранат")
async def cmd_refresh(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    await refresh_voice_presence()
    async with db_lock:
        data = load_db()
        _data, post_err = await upsert_status_messages(data)
    if post_err:
        await interaction.followup.send(
            embed=make_reply_embed(
                "⚠️  Не удалось обновить сообщения",
                post_err,
                color=COLOR_LIVE,
            ),
            ephemeral=True,
        )
        return
    await interaction.followup.send(
        embed=make_reply_embed(
            "🔄  Обновлено",
            "Сообщения **ЯВКА** и **ГРАНАТЫ** перерисованы.\nВойсы перепроверены.",
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


async def run_sheet_sync(
    *,
    force: bool = False,
    update_embeds: bool = True,
) -> tuple[bool, str, list[str]]:
    """
    Читает Excel и применяет к базе.
    force=True — всегда перечитать (игнор mtime/sig).
    update_embeds=False — только база (старт бота: embed один раз снаружи).
    """
    global _sheet_last_mtime, _sheet_last_sig, _sheet_last_error
    path = SHEET_PATH
    if not path.exists():
        msg = f"Файл не найден:\n`{path}`"
        _sheet_last_error = msg
        return False, msg, []

    lock = _sheet_lock_path(path)
    excel_open = lock.exists()

    try:
        mtime = path.stat().st_mtime
    except OSError as e:
        msg = f"Не прочитать файл: {e}"
        _sheet_last_error = msg
        return False, msg, []

    loop = asyncio.get_running_loop()
    try:
        entries, squad_names, tech_map = await loop.run_in_executor(
            None, read_squad_sheet, path
        )
    except PermissionError:
        msg = (
            f"Файл занят Excel — **сохрани (Ctrl+S)** или закрой таблицу:\n`{path}`"
        )
        _sheet_last_error = msg
        return False, msg, []
    except Exception as e:
        msg = f"Ошибка чтения Excel: `{e}`"
        _sheet_last_error = msg
        log(f"excel: {e}", "err")
        return False, msg, []

    sig = _entries_signature(entries)
    if (
        not force
        and _sheet_last_sig is not None
        and sig == _sheet_last_sig
        and mtime == _sheet_last_mtime
    ):
        hint = ""
        if excel_open:
            hint = " · Excel открыт — если правил, жми Ctrl+S"
        return True, f"на диске без изменений{hint}", []

    real: list[str] = []
    warns: list[str] = []
    data_after: dict[str, Any] | None = None
    async with db_lock:
        data = load_db()
        data, changes = apply_sheet_entries(data, entries, squad_names, tech_map)
        real = [c for c in changes if not str(c).startswith("⚠️")]
        warns = [c for c in changes if str(c).startswith("⚠️")]
        if real:
            save_db(data)
            data_after = data
        if update_embeds and (real or force):
            await upsert_status_messages(data, force=True)

    # tech = players (если Excel закрыт); иначе pending до следующего синка
    if data_after is not None:
        await sync_tech_sheet(data_after)
    elif _tech_pending:
        await sync_tech_sheet(None)

    _sheet_last_mtime = mtime
    _sheet_last_sig = sig
    _sheet_last_error = None

    open_note = ""
    if excel_open:
        open_note = " · Excel открыт — сохрани Ctrl+S, если правил"

    if real:
        return True, f"применено изменений: {len(real)}{open_note}", real + warns
    if warns:
        return True, f"файл ок, предупреждений: {len(warns)}{open_note}", warns
    return (
        True,
        f"файл прочитан, строк: {len(entries)}, без изменений{open_note}",
        [],
    )


@bot.tree.command(name="sheet_sync", description="Синхронизировать состав/отряды из Excel прямо сейчас")
async def cmd_sheet_sync(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    ok, msg, changes = await run_sheet_sync(force=True)
    if not ok:
        await interaction.followup.send(
            embed=make_reply_embed("❌  Sheet sync", msg, color=COLOR_ERR),
            ephemeral=True,
        )
        return

    # мини-сводка: что бот видит в Excel для пары контрольных ников
    try:
        entries, _n, _t = await asyncio.get_running_loop().run_in_executor(
            None, read_squad_sheet, SHEET_PATH
        )
        watch = {"itsgotime", "rickff"}
        seen = [
            f"`{e['game_nick']}` → отр.{e.get('squad')} слот {e.get('slot')}"
            for e in entries
            if (e.get("game_nick") or "").lower() in watch
        ]
        watch_txt = "\n".join(f"· {s}" for s in seen) if seen else "· (нет itsgotime/RickFF в таблице)"
    except Exception:
        watch_txt = "· (не удалось перечитать для сводки)"

    body = (
        f"**Файл:** `{SHEET_PATH.name}`\n"
        f"**Результат:** {msg}\n"
        f"\n**Сейчас в Excel (как видит бот):**\n{watch_txt}\n"
    )
    if changes:
        show = changes[:15]
        body += "\n**Изменения:**\n" + "\n".join(f"· {c}" for c in show)
        if len(changes) > 15:
            body += f"\n… и ещё {len(changes) - 15}"
    body += (
        f"\n\nАвто ~**{SHEET_SYNC_SECONDS}** сек. "
        f"После правки обязательно **Ctrl+S**, потом `/sheet_sync`."
    )
    await interaction.followup.send(
        embed=make_reply_embed("✅  Sheet sync" if ok else "❌  Sheet sync", body, color=COLOR_OK),
        ephemeral=True,
    )


@bot.tree.command(name="squad_list", description="Показать сетку отрядов (из базы после Excel)")
async def cmd_squad_list(interaction: discord.Interaction):
    data = load_db()
    players = data.get("players", {})
    by_squad: dict[Any, list[tuple[int, str, str]]] = {}
    no_squad: list[tuple[str, str]] = []

    for did, p in players.items():
        sq = p.get("squad")
        sl = p.get("slot")
        nick = p.get("game_nick") or "?"
        if sq is None:
            no_squad.append((did, nick))
            continue
        try:
            sl_i = int(sl) if sl is not None else 0
        except (TypeError, ValueError):
            sl_i = 0
        by_squad.setdefault(sq, []).append((sl_i, did, nick))

    lines: list[str] = []
    for sq in sorted(by_squad.keys(), key=lambda x: (isinstance(x, str), x)):
        lines.append(f"**{squad_label(data, sq)}**")
        for sl_i, did, nick in sorted(by_squad[sq], key=lambda t: (t[0], t[2].lower())):
            lines.append(f"`{sl_i:>2}.` `{nick}` | <@{did}>")
        lines.append("")
    if no_squad:
        lines.append("**Без отряда**")
        for did, nick in sorted(no_squad, key=lambda t: t[1].lower()):
            lines.append(f"· `{nick}` | <@{did}>")

    if not lines:
        text = "_пусто_"
    else:
        text = "\n".join(lines)

    embed = discord.Embed(
        title="🪖  ОТРЯДЫ",
        description=text[:4000],
        color=COLOR_INFO,
        timestamp=now_msk(),
    )
    embed.set_footer(text=f"Excel: {SHEET_PATH.name} · /sheet_sync")
    await interaction.response.send_message(embed=embed, ephemeral=True)


@bot.tree.command(name="sheet_path", description="Показать путь к Excel-таблице отрядов")
async def cmd_sheet_path(interaction: discord.Interaction):
    exists = SHEET_PATH.exists()
    err = f"\nПоследняя ошибка: {_sheet_last_error}" if _sheet_last_error else ""
    await interaction.response.send_message(
        embed=make_reply_embed(
            "📄  Excel отрядов",
            (
                f"**Путь:** `{SHEET_PATH}`\n"
                f"**Существует:** {'✅ да' if exists else '❌ нет'}\n"
                f"**Автосинк:** каждые {SHEET_SYNC_SECONDS} сек\n"
                f"Задай `SHEET_PATH` в `.env` если файл в OneDrive.\n"
                f"{err}"
            ),
            color=COLOR_OK if exists else COLOR_ERR,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="scan_now", description="Запустить скан гранат вручную")
@app_commands.describe(step="Какой этап просканировать")
@app_commands.choices(
    step=[
        app_commands.Choice(name="База (чт–сб 20:05 / вс 19:00)", value="20:00"),
        app_commands.Choice(name="I (чт–сб 20:25 / вс 19:20)", value="20:25"),
        app_commands.Choice(name="II (чт–сб 20:50 / вс 19:40)", value="20:50"),
        app_commands.Choice(name="III (чт–сб 21:20 финал / вс 20:00)", value="21:20"),
        app_commands.Choice(name="IV финал вс 20:20", value="21:40"),
    ]
)
async def cmd_scan_now(interaction: discord.Interaction, step: app_commands.Choice[str]):
    actor = await resolve_member(interaction)
    if not can_manage_kv(actor):
        await interaction.response.send_message(embed=deny_embed(), ephemeral=True)
        return
    await interaction.response.defer(ephemeral=True)
    prev_map = {
        "20:00": None,
        "20:25": "20:00",
        "20:50": "20:25",
        "21:20": "20:50",
        "21:40": "21:20",
    }
    await run_grenade_step(step.value, prev_map[step.value])
    await interaction.followup.send(
        embed=make_reply_embed(
            "💣  Скан выполнен",
            f"**{STEP_TITLES.get(step.value, step.value)}** (`{step.value}`) — готово.\n"
            f"Смотри канал логов и таблицу **ГРАНАТЫ**.\n"
            f"_Если API fail — этап не закрыт, можно повторить._",
            color=COLOR_GRENADES,
        ),
        ephemeral=True,
    )


@bot.tree.command(
    name="deletegren",
    description="Очистить таблицу ГРАНАТЫ (ники, цифры и дату)",
)
async def cmd_deletegren(interaction: discord.Interaction):
    actor = await resolve_member(interaction)
    if not can_manage_kv(actor):
        await interaction.response.send_message(embed=deny_embed(), ephemeral=True)
        return
    async with db_lock:
        data = load_db()
        backup_db()
        old_label = format_kv_date_label(data) if (
            data.get("grenade_history") or data.get("grenade_date")
        ) else None
        data["grenade_history"] = {}
        data.pop("combat_history", None)
        data["skipped_grenade_steps"] = []
        data["last_grenade_step"] = None
        data["grenade_date"] = None
        # дату/таблицу убрали; финал КВ по явке не трогаем
        save_db(data)
        data_out = data
    await upsert_status_messages(data_out)

    extra = f"Было: **{old_label}**\n" if old_label else ""
    await interaction.response.send_message(
        embed=make_reply_embed(
            "🧹  Гранаты очищены",
            (
                f"{extra}"
                f"Таблица **ГРАНАТЫ** пустая — без ников, цифр и даты.\n"
                f"Явка не тронута."
            ),
            color=COLOR_OK,
        ),
        ephemeral=True,
    )


@bot.tree.command(name="reset_session", description="Сбросить явку и гранаты (новая сессия)")
async def cmd_reset_session(interaction: discord.Interaction):
    actor = await resolve_member(interaction)
    if not can_manage_kv(actor):
        await interaction.response.send_message(embed=deny_embed(), ephemeral=True)
        return
    async with db_lock:
        data = load_db()
        backup_db()
        data = await ensure_session_reset(data, force=True)
        # force сам чистит; если ещё до 19:30 — force не откроет сессию, чистим вручную
        if not data.get("kv_session_active"):
            data["grenade_history"] = {}
            data.pop("combat_history", None)
            data["skipped_grenade_steps"] = []
            data["last_grenade_step"] = None
            data["kv_finished"] = False
            data["grenade_date"] = today_msk_str()
            for p in data.get("players", {}).values():
                p["came"] = False
                p["in_voice"] = False
        data.pop("combat_history", None)
        data["skipped_grenade_steps"] = []
        # полный сброс сессии — карты тоже
        clear_kv_maps(data)
        guild = interaction.guild
        if guild and is_kv_live(data):
            voice_ids = collect_voice_member_ids(guild)
            for did_str, p in data["players"].items():
                if int(did_str) in voice_ids:
                    p["in_voice"] = True
                    if can_mark_came(data):
                        p["came"] = True
        save_db(data)
        data_out = data

    await upsert_status_messages(data_out, force=True)
    await interaction.response.send_message(
        embed=make_reply_embed(
            "♻️  Сессия сброшена",
            (
                f"Дата: **{format_kv_date_label(load_db())}**\n"
                f"· Явка обнулена\n"
                f"· История гранат очищена\n"
                f"· Карты этапов сброшены\n"
                f"· В окне КВ кто в войсе — снова ✅\n"
                f"· backup: `players.json.bak`"
            ),
            color=COLOR_INFO,
        ),
    )


# ---------------------------------------------------------------------------
# Фоновые задачи
# ---------------------------------------------------------------------------
@tasks.loop(seconds=15)
async def voice_watch_loop():
    try:
        await refresh_voice_presence()
    except Exception as e:
        log(f"войс: {e}", "err")


@tasks.loop(seconds=SHEET_SYNC_SECONDS)
async def sheet_sync_loop():
    """Авточтение Excel (ники / отряды / слоты)."""
    try:
        ok, msg, changes = await run_sheet_sync(force=False)
        if not ok:
            log(f"excel · {msg}", "err")
        elif changes and any(not str(c).startswith("⚠️") for c in changes):
            n = sum(1 for c in changes if not str(c).startswith("⚠️"))
            log(f"excel · обновлено {n}", "ok")
    except Exception as e:
        log(f"excel: {e}", "err")


@tasks.loop(seconds=20)
async def schedule_loop():
    """
    Расписание по дню (МСК):
    чт–сб: 19:30 явка → 20:05 база → 20:25/50 → 21:20
    вс:    18:30 явка → 19:00 база → 19:20/40 → 20:20
    """
    try:
        now = now_msk()
        if not is_cw_day(now):
            return

        t = now.time().replace(second=0, microsecond=0)
        minute_key = now.strftime("%Y-%m-%d %H:%M")
        start_t = kv_start(now)
        absent_t = kv_absent_dm(now)
        steps = kv_grenade_steps(now)

        if not hasattr(schedule_loop, "_done"):
            schedule_loop._done = set()  # type: ignore[attr-defined]
        done: set = schedule_loop._done  # type: ignore[attr-defined]

        # Старт явки ИЛИ догон, если бот включили уже после
        start_key = f"{now.strftime('%Y-%m-%d')}:start"
        if start_key not in done and t >= start_t:
            opened = False
            async with db_lock:
                data = load_db()
                already_live = (
                    data.get("session_date") == today_msk_str()
                    and data.get("kv_session_active")
                    and not is_kv_finished(data)
                )
                if not already_live:
                    data = await ensure_session_reset(
                        data, force=(t == start_t)
                    )
                    if data.get("kv_session_active") and not is_kv_finished(data):
                        opened = True
                        await upsert_status_messages(data, force=True)
                done.add(start_key)
            if opened:
                await refresh_voice_presence()
                log(f"явка открыта · {now.strftime('%H:%M')} МСК", "kv")

        # ЛС неявившимся
        if t == absent_t and f"{minute_key}:absent_dm" not in done:
            done.add(f"{minute_key}:absent_dm")
            await refresh_voice_presence()
            guild = bot.get_guild(get_guild_id()) if get_guild_id() else None
            result = await dm_never_came_players(guild, force_resend=False)
            log(
                f"ЛС неявившимся · ok {result['ok']} · fail {result['fail']}"
                f" · skip замены {result.get('skip_subs', 0)}",
                "kv",
            )

        # Этапы грен (без фейк-догона I/II — API не знает прошлое)
        today = today_msk_str()
        for step_name, step_time, prev in steps:
            day_step_key = f"{today}:{step_name}"
            if day_step_key in done:
                continue
            if t < step_time:
                break

            data_chk = load_db()
            if _step_already_done(data_chk.get("last_grenade_step"), step_name):
                done.add(day_step_key)
                continue

            late = _minutes_late(t, step_time)
            # mid (I/II…): окно 0–1 мин после времени; иначе skip + forward-fill
            mid_steps = {s[0] for s in steps[1:-1]}
            if step_name in mid_steps:
                if late < 0:
                    break
                if late > 1:
                    done.add(day_step_key)
                    await fill_skipped_grenade_step(step_name)
                    continue
            else:
                # база / финал: догон если бот включили позже
                if late < 0:
                    break

            # НЕ ставим done до успеха: иначе API fail = этап навсегда пропущен
            await refresh_voice_presence()
            before_last = load_db().get("last_grenade_step")
            await run_grenade_step(step_name, prev)
            after = load_db()
            after_last = after.get("last_grenade_step")
            if _step_already_done(after_last, step_name):
                done.add(day_step_key)
            else:
                # fail — пробуем снова на следующих тиках (до ~5 мин после времени)
                if late > 5:
                    done.add(day_step_key)
                    log(
                        f"этап {step_name} сдан без данных (API fail >5 мин) · "
                        f"ручной /scan_now",
                        "err",
                    )
                elif before_last == after_last:
                    log(
                        f"этап {step_name} · API не прошёл · повтор через ~20с",
                        "warn",
                    )

        if len(done) > 80:
            schedule_loop._done = {k for k in done if k.startswith(today)}  # type: ignore

    except Exception as e:
        log(f"расписание: {e}", "err")


@voice_watch_loop.before_loop
async def before_voice():
    await bot.wait_until_ready()


@sheet_sync_loop.before_loop
async def before_sheet():
    await bot.wait_until_ready()
    # первый проход уже в on_ready — цикл ждёт интервал, без двойного PATCH/429
    await asyncio.sleep(SHEET_SYNC_SECONDS)


@schedule_loop.before_loop
async def before_schedule():
    await bot.wait_until_ready()


@bot.tree.error
async def on_app_command_error(
    interaction: discord.Interaction,
    error: app_commands.AppCommandError,
):
    log(f"/{interaction.command}: {error}", "err")
    try:
        msg = f"Ошибка команды: `{error}`"
        if interaction.response.is_done():
            await interaction.followup.send(msg, ephemeral=True)
        else:
            await interaction.response.send_message(msg, ephemeral=True)
    except Exception:
        pass


@bot.event
async def on_ready():
    for name in ("add", "remove", "list"):
        cmd = bot.tree.get_command(name)
        if cmd is not None:
            try:
                cmd.default_permissions = None
            except Exception:
                pass

    if bot.guilds:
        for g in bot.guilds:
            try:
                bot.tree.copy_global_to(guild=g)
                await bot.tree.sync(guild=g)
            except Exception as e:
                log(f"sync «{g.name}»: {e}", "err")
    else:
        await bot.tree.sync()

    if not voice_watch_loop.is_running():
        voice_watch_loop.start()
    if not schedule_loop.is_running():
        schedule_loop.start()
    if not sheet_sync_loop.is_running():
        sheet_sync_loop.start()

    await run_sheet_sync(force=True, update_embeds=False)

    n = now_msk()
    async with db_lock:
        data = load_db()
        mid_kv = (
            bool(data.get("kv_session_active"))
            and data.get("session_date") == today_msk_str()
            and not is_kv_finished(data)
        )
        if not mid_kv:
            data = reset_attendance_only(data)
            save_db(data)
        data = await ensure_session_reset(data)
        fin = is_kv_finished(data)
        live = is_kv_live(data)
        if get_log_channel_id(data):
            await upsert_status_messages(data, force=True)
        else:
            log("нет канала логов · /setup", "warn")

    await refresh_voice_presence()
    data = load_db()
    fin = is_kv_finished(data)
    live = is_kv_live(data)
    sess = data.get("session_date")
    today = today_msk_str()

    # статус баннера = как в embed, без «finished» с прошлого дня
    if live:
        status = "КВ live"
    elif not is_cw_day(n):
        status = "вне КВ"
    elif fin and sess == today:
        if n.time() < kv_final_time(n):
            status = "finished · /reset_session?"
        else:
            status = "КВ сегодня закрыт"
    elif n.time() < kv_start(n):
        status = f"ожидание с {kv_start(n).strftime('%H:%M')}"
    else:
        status = "день КВ"

    log_banner(
        [
            f"{bot.user}",
            f"{WEEKDAYS_RU[n.weekday()]} {n.strftime('%H:%M')} МСК · {status}",
            f"region {REGION} · войсов {len(get_voice_channel_ids(data))}",
            f"excel {SHEET_PATH.name}",
        ]
    )


@bot.event
async def on_guild_join(guild: discord.Guild):
    try:
        bot.tree.copy_global_to(guild=guild)
        await bot.tree.sync(guild=guild)
        log(f"сервер «{guild.name}»", "ok")
    except Exception as e:
        log(f"join sync: {e}", "err")


@bot.event
async def on_voice_state_update(
    member: discord.Member,
    before: discord.VoiceState,
    after: discord.VoiceState,
):
    if member.bot:
        return

    need_upsert = False
    async with db_lock:
        data = load_db()
        guild_id = get_guild_id(data)
        if guild_id and member.guild.id != guild_id:
            return

        relevant = set(get_voice_channel_ids(data))
        if not relevant:
            return

        before_id = before.channel.id if before.channel else None
        after_id = after.channel.id if after.channel else None
        if before_id not in relevant and after_id not in relevant:
            return

        key = str(member.id)
        if key not in data.get("players", {}):
            return

        data = await ensure_session_reset(data)
        live = is_kv_live(data)
        p = data["players"][key]
        p["discord_name"] = member.display_name
        in_target = bool(after_id in relevant) if after_id else False

        if not live:
            if p.get("in_voice"):
                p["in_voice"] = False
                save_db(data)
                need_upsert = True
            # fall through to upsert outside lock
        else:
            p["in_voice"] = in_target
            if can_mark_came(data) and in_target and not p.get("came"):
                p["came"] = True
                log(f"явка · {p.get('game_nick')}", "ok")
            save_db(data)
            need_upsert = True

    if need_upsert:
        await upsert_status_messages(None, force=False, parts=("online",))


def main():
    if not DISCORD_TOKEN:
        raise SystemExit("  ✗  нет DISCORD_TOKEN в .env")
    if not CLIENT_SECRET:
        log("STALCRAFT_CLIENT_SECRET пуст — грены не будут сканиться", "warn")

    _orig_close = bot.close

    async def _close_with_http() -> None:
        await close_http()
        await _orig_close()

    bot.close = _close_with_http  # type: ignore[method-assign]
    bot.run(DISCORD_TOKEN)


if __name__ == "__main__":
    main()
